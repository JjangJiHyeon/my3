"""
FastAPI backend for the multi-format document parser.

Configuration via environment variables:
  DOC_DIR     – directory containing source documents  (default: ./documents)
  RESULT_DIR  – directory for cached parse results      (default: ./parsed_results)
  CLEAN_GENERATED_ON_STARTUP – if "1", delete generated artifacts on server start
       (parsed_results JSON/review/previews, exports/, root *_llm_*.json, etc.)
       Default "0". Same as legacy DEV_CLEAR_RESULTS_ON_START if the latter is "1".
"""

from __future__ import annotations

import hashlib
import json
import logging
import os
import shutil
import subprocess
import sys
import threading
import uuid
from collections import Counter
from contextlib import asynccontextmanager
from datetime import datetime
from html import escape as html_escape
from pathlib import Path, PureWindowsPath
from typing import Any
from urllib import error, request

from fastapi import File, FastAPI, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles

from app_support.artifact_cleanup import cleanup_generated_artifacts
from app_support.export_to_gpt import write_root_llm_exports
from parsers import parse_document, SUPPORTED_EXTENSIONS, PARSER_VERSION
from app_support.review_export import save_all_parse_outputs

logging.basicConfig(level=logging.INFO, format="%(levelname)s  %(name)s  %(message)s")
logger = logging.getLogger(__name__)

DOC_DIR: str = os.getenv(
    "DOC_DIR",
    os.path.join(os.path.dirname(__file__), "documents"),
)
RESULT_DIR: str = os.getenv(
    "RESULT_DIR",
    os.path.join(os.path.dirname(__file__), "parsed_results"),
)
PROJECT_ROOT = Path(__file__).resolve().parent
CLEAN_GENERATED_ON_STARTUP: bool = (
    os.getenv("CLEAN_GENERATED_ON_STARTUP", os.getenv("DEV_CLEAR_RESULTS_ON_START", "0")) == "1"
)
CLEAR_VIEWER_CACHE_ON_STARTUP: bool = os.getenv("CLEAR_VIEWER_CACHE_ON_STARTUP", "0") == "1"
RAG_API_BASE_URL = os.getenv("RAG_API_BASE_URL", "http://127.0.0.1:8000").rstrip("/")
ENABLE_HWP_PDF_SIDECAR: bool = os.getenv("ENABLE_HWP_PDF_SIDECAR", "0") == "1"
LIBREOFFICE_BIN: str = os.getenv("LIBREOFFICE_BIN", "").strip()
HWP_CONVERTER_SCRIPT = PROJECT_ROOT / "convert_hwp.py"
DEFAULT_HWP_PROFILE_DIR = PROJECT_ROOT / "_tmp_lo_unopkg_profile"

os.makedirs(DOC_DIR, exist_ok=True)
os.makedirs(RESULT_DIR, exist_ok=True)

if not os.getenv("HWP_LIBREOFFICE_PROFILE", "").strip():
    DEFAULT_HWP_PROFILE_DIR.mkdir(parents=True, exist_ok=True)
    os.environ["HWP_LIBREOFFICE_PROFILE"] = str(DEFAULT_HWP_PROFILE_DIR)


def _clear_viewer_cache_files() -> None:
    previews_dir = Path(RESULT_DIR) / "previews"
    if previews_dir.exists():
        try:
            shutil.rmtree(previews_dir)
            logger.info("Cleared viewer cache directory %s", previews_dir)
        except OSError as exc:
            logger.warning("Could not clear viewer cache directory %s: %s", previews_dir, exc)
    try:
        previews_dir.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        logger.warning("Could not recreate viewer cache directory %s: %s", previews_dir, exc)


@asynccontextmanager
async def lifespan(application: FastAPI):
    # Tell the PDF parser where to store preview images
    try:
        from parsers.pdf_parser import set_result_dir
        set_result_dir(RESULT_DIR)
    except Exception as exc:
        logger.warning("Could not set RESULT_DIR on pdf_parser: %s", exc)

    if CLEAN_GENERATED_ON_STARTUP:
        logger.info("CLEAN_GENERATED_ON_STARTUP: clearing generated artifacts")
        cleanup_generated_artifacts(
            PROJECT_ROOT,
            parsed_results_dir=Path(RESULT_DIR),
        )
        parsed_cache.clear()
    elif CLEAR_VIEWER_CACHE_ON_STARTUP:
        logger.info("CLEAR_VIEWER_CACHE_ON_STARTUP: clearing viewer cache artifacts")
        _clear_viewer_cache_files()

    _warm_cache()
    hwp_sidecar_failure_cache.clear()
    logger.info("DOC_DIR    = %s", DOC_DIR)
    logger.info("RESULT_DIR = %s", RESULT_DIR)
    logger.info("CLEAN_GENERATED_ON_STARTUP = %s", CLEAN_GENERATED_ON_STARTUP)
    logger.info("CLEAR_VIEWER_CACHE_ON_STARTUP = %s", CLEAR_VIEWER_CACHE_ON_STARTUP)
    yield


app = FastAPI(title="Document Parser", lifespan=lifespan)
app.mount("/static", StaticFiles(directory=PROJECT_ROOT / "static"), name="static")

parsed_cache: dict[str, dict[str, Any]] = {}
upload_jobs: dict[str, dict[str, Any]] = {}
upload_jobs_lock = threading.Lock()
index_build_lock = threading.Lock()
rag_response_cache_lock = threading.Lock()
hwp_sidecar_failure_cache: set[str] = set()
summary_response_cache: dict[tuple[str, str, str], dict[str, Any]] = {}
qa_response_cache: dict[tuple[str, str, str, str], dict[str, Any]] = {}


def _summary_cache_key(payload: dict[str, Any]) -> tuple[str, str, str]:
    return (
        str(payload.get("filename") or "").strip(),
        str(payload.get("strategy_name") or "").strip(),
        str(payload.get("top_k") or "").strip(),
    )


def _qa_cache_key(payload: dict[str, Any]) -> tuple[str, str, str, str]:
    return (
        str(payload.get("filename_filter") or "").strip(),
        str(payload.get("query") or "").strip(),
        str(payload.get("strategy_name") or "").strip(),
        str(payload.get("top_k") or "").strip(),
    )


def _invalidate_rag_response_cache(*, filename: str | None = None) -> None:
    with rag_response_cache_lock:
        if not filename:
            summary_response_cache.clear()
            qa_response_cache.clear()
            return

        normalized = filename.strip()
        for key in list(summary_response_cache.keys()):
            if key[0] == normalized:
                summary_response_cache.pop(key, None)
        for key in list(qa_response_cache.keys()):
            if key[0] == normalized:
                qa_response_cache.pop(key, None)


# ── helpers ──────────────────────────────────────────────────────────

def _doc_id(filepath: str) -> str:
    return hashlib.md5(filepath.encode("utf-8")).hexdigest()


def _result_path(doc_id: str) -> Path:
    return Path(RESULT_DIR) / f"{doc_id}.json"


def _preview_path(doc_id: str, page_num: int) -> Path:
    return Path(RESULT_DIR) / "previews" / doc_id / f"page_{page_num}.png"


def _preview_dir(doc_id: str) -> Path:
    return Path(RESULT_DIR) / "previews" / doc_id


def _stitched_preview_path(doc_id: str) -> Path:
    return _preview_dir(doc_id) / "stitched_horizontal.png"


def _sidecar_pdf_path(doc_id: str) -> Path:
    return _preview_dir(doc_id) / "source.pdf"


def _existing_sidecar_pdf(doc_id: str) -> Path | None:
    sidecar_path = _sidecar_pdf_path(doc_id)
    if sidecar_path.is_file():
        return sidecar_path
    preview_dir = _preview_dir(doc_id)
    if not preview_dir.exists():
        return None
    pdf_candidates = sorted(
        (candidate for candidate in preview_dir.glob("*.pdf") if candidate.is_file()),
        key=lambda candidate: candidate.stat().st_mtime,
        reverse=True,
    )
    return pdf_candidates[0] if pdf_candidates else None


def _parsed_doc_id(parsed: dict[str, Any]) -> str:
    doc_id = str(parsed.get("id") or "").strip()
    if doc_id:
        return doc_id
    filepath = str(parsed.get("filepath") or "").strip()
    return _doc_id(filepath) if filepath else ""


def _sidecar_pdf_enabled(file_type: str) -> bool:
    normalized = str(file_type or "").lower()
    if normalized in {"doc", "docx", "xls", "xlsx", "hwp"}:
        return True
    return False


def _viewer_sidecar_generation_allowed(file_type: str) -> bool:
    return str(file_type or "").lower() in {"doc", "docx", "xls", "xlsx", "hwp"}


def _resolve_libreoffice_bin() -> Path | None:
    candidates = [
        LIBREOFFICE_BIN,
        shutil.which("soffice.exe"),
        shutil.which("soffice"),
        str(Path(os.environ.get("PROGRAMFILES", r"C:\Program Files")) / "LibreOffice" / "program" / "soffice.exe"),
        str(
            Path(os.environ.get("PROGRAMFILES(X86)", r"C:\Program Files (x86)"))
            / "LibreOffice"
            / "program"
            / "soffice.exe"
        ),
    ]
    for candidate in candidates:
        if not candidate:
            continue
        candidate_path = Path(str(candidate))
        if candidate_path.is_file():
            if candidate_path.suffix.lower() == ".com":
                exe_candidate = candidate_path.with_suffix(".exe")
                if exe_candidate.is_file():
                    return exe_candidate
            return candidate_path
        resolved = shutil.which(str(candidate))
        if resolved:
            resolved_path = Path(resolved)
            if resolved_path.suffix.lower() == ".com":
                exe_candidate = resolved_path.with_suffix(".exe")
                if exe_candidate.is_file():
                    return exe_candidate
            return resolved_path
    return None


def _clear_generated_previews(doc_id: str) -> None:
    preview_dir = _preview_dir(doc_id)
    if not preview_dir.exists():
        return
    for path in preview_dir.glob("page_*.png"):
        try:
            path.unlink()
        except FileNotFoundError:
            continue
        except Exception as exc:
            logger.warning("Could not remove stale preview %s: %s", path, exc)
    stitched_path = _stitched_preview_path(doc_id)
    try:
        if stitched_path.exists():
            stitched_path.unlink()
    except FileNotFoundError:
        pass
    except Exception as exc:
        logger.warning("Could not remove stitched preview %s: %s", stitched_path, exc)


def _libreoffice_run_environment(soffice_path: Path) -> dict[str, str]:
    env = os.environ.copy()
    for key in (
        "PYTHONHOME",
        "PYTHONPATH",
        "PYTHONEXECUTABLE",
        "VIRTUAL_ENV",
        "CONDA_PREFIX",
        "CONDA_DEFAULT_ENV",
    ):
        env.pop(key, None)
    env["PATH"] = str(soffice_path.parent) + os.pathsep + env.get("PATH", "")
    return env


def _libreoffice_subprocess_kwargs() -> dict[str, Any]:
    kwargs: dict[str, Any] = {}
    if os.name == "nt":
        startupinfo = subprocess.STARTUPINFO()
        startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
        kwargs["startupinfo"] = startupinfo
        kwargs["creationflags"] = getattr(subprocess, "CREATE_NO_WINDOW", 0)
    return kwargs


def _maybe_create_hwp_sidecar_pdf(parsed: dict[str, Any]) -> Path | None:
    file_type = str(parsed.get("file_type") or "").lower()
    if file_type != "hwp":
        return None

    filepath = str(parsed.get("filepath") or "").strip()
    source_path = Path(filepath) if filepath else None
    if source_path is None or not source_path.is_file():
        return None

    doc_id = _parsed_doc_id(parsed)
    if not doc_id:
        return None

    existing_pdf = _existing_sidecar_pdf(doc_id)
    try:
        if existing_pdf is not None and existing_pdf.stat().st_mtime >= source_path.stat().st_mtime:
            return existing_pdf
    except Exception:
        pass

    if doc_id in hwp_sidecar_failure_cache:
        return None

    if not HWP_CONVERTER_SCRIPT.is_file():
        logger.warning("HWP converter script not found: %s", HWP_CONVERTER_SCRIPT)
        return None

    preview_dir = _preview_dir(doc_id)
    preview_dir.mkdir(parents=True, exist_ok=True)
    legacy_failure_marker = preview_dir / ".sidecar_failed"
    try:
        if legacy_failure_marker.exists():
            legacy_failure_marker.unlink()
    except Exception:
        pass
    output_path = _sidecar_pdf_path(doc_id)

    command = [
        sys.executable,
        str(HWP_CONVERTER_SCRIPT),
        str(source_path.resolve()),
        str(output_path.resolve()),
    ]
    try:
        completed = subprocess.run(
            command,
            cwd=str(PROJECT_ROOT),
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=180,
            check=False,
            **_libreoffice_subprocess_kwargs(),
        )
    except Exception as exc:
        logger.warning("HWP sidecar conversion failed for %s: %s", source_path.name, exc)
        return None

    if completed.returncode != 0 or not output_path.is_file():
        detail = (completed.stderr or completed.stdout or "").strip()
        hwp_sidecar_failure_cache.add(doc_id)
        logger.warning(
            "HWP sidecar conversion returned %s for %s: %s",
            completed.returncode,
            source_path.name,
            detail,
        )
        return None

    hwp_sidecar_failure_cache.discard(doc_id)
    _clear_generated_previews(doc_id)
    return output_path


def _maybe_create_sidecar_pdf(parsed: dict[str, Any]) -> Path | None:
    file_type = str(parsed.get("file_type") or "").lower()
    if file_type == "hwp":
        return _maybe_create_hwp_sidecar_pdf(parsed)
    if not _sidecar_pdf_enabled(file_type):
        return None

    filepath = str(parsed.get("filepath") or "").strip()
    source_path = Path(filepath) if filepath else None
    if source_path is None or not source_path.is_file():
        return None

    doc_id = _parsed_doc_id(parsed)
    if not doc_id:
        return None

    sidecar_path = _sidecar_pdf_path(doc_id)
    existing_pdf = _existing_sidecar_pdf(doc_id)
    try:
        if existing_pdf is not None and existing_pdf.stat().st_mtime >= source_path.stat().st_mtime:
            return existing_pdf
    except Exception:
        pass

    soffice_path = _resolve_libreoffice_bin()
    if soffice_path is None:
        logger.warning("LibreOffice headless executable not found for sidecar conversion: %s", source_path.name)
        return None

    preview_dir = _preview_dir(doc_id)
    preview_dir.mkdir(parents=True, exist_ok=True)
    profile_dir = preview_dir / ".lo_profile"
    profile_dir.mkdir(parents=True, exist_ok=True)

    produced_pdf = preview_dir / f"{source_path.stem}.pdf"
    command = [
        str(soffice_path),
        "--headless",
        "--invisible",
        "--nologo",
        "--nodefault",
        "--nolockcheck",
        "--norestore",
        f"-env:UserInstallation={profile_dir.resolve().as_uri()}",
        "--convert-to",
        "pdf",
        "--outdir",
        str(preview_dir.resolve()),
        str(source_path.resolve()),
    ]
    try:
        completed = subprocess.run(
            command,
            cwd=str(soffice_path.parent),
            env=_libreoffice_run_environment(soffice_path),
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=180,
            check=False,
            **_libreoffice_subprocess_kwargs(),
        )
    except Exception as exc:
        logger.warning("LibreOffice sidecar conversion failed for %s: %s", source_path.name, exc)
        return None

    if not produced_pdf.is_file():
        pdf_candidates = sorted(
            (
                candidate
                for candidate in preview_dir.glob("*.pdf")
                if candidate.name.lower() != sidecar_path.name.lower()
            ),
            key=lambda candidate: candidate.stat().st_mtime,
            reverse=True,
        )
        if pdf_candidates:
            produced_pdf = pdf_candidates[0]

    if not produced_pdf.is_file():
        detail = (completed.stderr or completed.stdout or "").strip()
        logger.warning("LibreOffice sidecar conversion did not produce a PDF for %s", source_path.name)
        if completed.returncode != 0 or detail:
            logger.warning(
                "LibreOffice sidecar conversion returned %s for %s: %s",
                completed.returncode,
                source_path.name,
                detail,
            )
        return None

    final_pdf_path = produced_pdf
    try:
        if produced_pdf.resolve() != sidecar_path.resolve():
            if sidecar_path.exists():
                sidecar_path.unlink()
            shutil.move(str(produced_pdf), str(sidecar_path))
            final_pdf_path = sidecar_path
    except Exception as exc:
        logger.warning("Could not normalize sidecar PDF name for %s: %s", source_path.name, exc)

    _clear_generated_previews(doc_id)
    return final_pdf_path if final_pdf_path.is_file() else None


def _viewer_pdf_source_path(
    parsed: dict[str, Any],
    *,
    generate_if_missing: bool = False,
) -> Path | None:
    file_type = str(parsed.get("file_type") or "").lower()
    filepath = str(parsed.get("filepath") or "").strip()

    if file_type == "pdf":
        pdf_path = Path(filepath) if filepath else None
        return pdf_path if pdf_path and pdf_path.is_file() else None

    doc_id = _parsed_doc_id(parsed)
    if not doc_id:
        return None

    sidecar_path = _existing_sidecar_pdf(doc_id)
    source_path = Path(filepath) if filepath else None
    if sidecar_path is not None and sidecar_path.is_file():
        if source_path is None or not source_path.is_file():
            return sidecar_path
        try:
            if sidecar_path.stat().st_mtime >= source_path.stat().st_mtime:
                return sidecar_path
        except Exception:
            return sidecar_path

    if not generate_if_missing:
        return None
    return _maybe_create_sidecar_pdf(parsed)


def _pdf_page_count(pdf_path: Path) -> int | None:
    try:
        import fitz  # PyMuPDF
    except Exception as exc:
        logger.warning("Could not import PyMuPDF for page count: %s", exc)
        return None

    try:
        with fitz.open(str(pdf_path)) as document:
            return int(document.page_count)
    except Exception as exc:
        logger.warning("Could not read PDF page count for %s: %s", pdf_path, exc)
        return None


def _ensure_pdf_preview(doc_id: str, page_num: int) -> Path | None:
    preview_path = _preview_path(doc_id, page_num)
    cached = _load_result(doc_id) or parsed_cache.get(doc_id)
    if preview_path.is_file():
        if not cached:
            return preview_path
        pdf_source = _viewer_pdf_source_path(cached, generate_if_missing=False)
        if pdf_source is None or not pdf_source.is_file():
            return preview_path
        try:
            if preview_path.stat().st_mtime >= pdf_source.stat().st_mtime:
                return preview_path
        except Exception:
            return preview_path

    if not cached:
        return None

    pdf_source = _viewer_pdf_source_path(
        cached,
        generate_if_missing=_viewer_sidecar_generation_allowed(str(cached.get("file_type") or "")),
    )
    if pdf_source is None or not pdf_source.is_file():
        return None

    try:
        import fitz  # PyMuPDF
    except Exception as exc:
        logger.warning("Could not import PyMuPDF for preview fallback: %s", exc)
        return None

    try:
        with fitz.open(str(pdf_source)) as document:
            if page_num < 1 or page_num > document.page_count:
                return None
            page = document.load_page(page_num - 1)
            pixmap = page.get_pixmap(matrix=fitz.Matrix(1.5, 1.5))
            preview_path.parent.mkdir(parents=True, exist_ok=True)
            pixmap.save(str(preview_path))
            return preview_path
    except Exception as exc:
        logger.warning("Failed to generate preview for doc=%s page=%s: %s", doc_id, page_num, exc)
        return None


def _preview_content_bbox(image: Any) -> tuple[int, int, int, int] | None:
    try:
        from PIL import ImageOps

        gray = ImageOps.grayscale(image.convert("RGB"))
        mask = gray.point(lambda value: 255 if value < 245 else 0)
        return mask.getbbox()
    except Exception:
        return None


def _ensure_spreadsheet_stitched_preview(doc_id: str) -> Path | None:
    cached = _load_result(doc_id) or parsed_cache.get(doc_id)
    if not cached:
        return None

    file_type = str(cached.get("file_type") or "").lower()
    if file_type not in {"xls", "xlsx"}:
        return None

    pdf_source = _viewer_pdf_source_path(cached, generate_if_missing=True)
    if pdf_source is None or not pdf_source.is_file():
        return None

    page_count = _pdf_page_count(pdf_source) or len(cached.get("pages") or []) or 1
    if page_count <= 1:
        return _ensure_pdf_preview(doc_id, 1)

    stitched_path = _stitched_preview_path(doc_id)
    preview_paths: list[Path] = []
    for page_num in range(1, page_count + 1):
        preview_path = _ensure_pdf_preview(doc_id, page_num)
        if preview_path is None or not preview_path.is_file():
            return None
        preview_paths.append(preview_path)

    try:
        newest_source_mtime = max(path.stat().st_mtime for path in preview_paths)
        if stitched_path.is_file() and stitched_path.stat().st_mtime >= newest_source_mtime:
            return stitched_path
    except Exception:
        pass

    try:
        from PIL import Image
    except Exception as exc:
        logger.warning("Could not import Pillow for stitched spreadsheet preview: %s", exc)
        return None

    try:
        opened_images = [Image.open(path).convert("RGB") for path in preview_paths]
        bboxes = [_preview_content_bbox(image) for image in opened_images]
        valid_bboxes = [bbox for bbox in bboxes if bbox is not None]
        if valid_bboxes:
            top = min(bbox[1] for bbox in valid_bboxes)
            bottom = max(bbox[3] for bbox in valid_bboxes)
        else:
            top = 0
            bottom = max(image.height for image in opened_images)

        cropped_images = []
        for image, bbox in zip(opened_images, bboxes):
            if bbox is None:
                left, right = 0, image.width
            else:
                left, right = bbox[0], bbox[2]
            crop_top = max(0, min(top, image.height))
            crop_bottom = max(crop_top + 1, min(bottom, image.height))
            crop_left = max(0, min(left, image.width - 1))
            crop_right = max(crop_left + 1, min(right, image.width))
            cropped_images.append(image.crop((crop_left, crop_top, crop_right, crop_bottom)))

        total_width = sum(image.width for image in cropped_images)
        max_height = max(image.height for image in cropped_images)
        canvas = Image.new("RGB", (total_width, max_height), "white")
        offset_x = 0
        for image in cropped_images:
            canvas.paste(image, (offset_x, 0))
            offset_x += image.width

        stitched_path.parent.mkdir(parents=True, exist_ok=True)
        canvas.save(stitched_path, format="PNG")
        return stitched_path
    except Exception as exc:
        logger.warning("Failed to build stitched spreadsheet preview for doc=%s: %s", doc_id, exc)
        return None


def _viewer_html(text: Any) -> str:
    return html_escape("" if text is None else str(text), quote=True)


def _viewer_multiline_html(text: Any) -> str:
    lines = str(text or "").splitlines() or [""]
    return "<br>".join(_viewer_html(line) for line in lines)


def _viewer_trim_row(cells: list[str]) -> list[str]:
    trimmed = list(cells)
    while trimmed and not str(trimmed[-1] or "").strip():
        trimmed.pop()
    return trimmed


def _viewer_table_rows_from_pipe_text(text: Any) -> list[list[str]]:
    rows: list[list[str]] = []
    for raw_line in str(text or "").splitlines():
        line = raw_line.strip()
        if not line or "|" not in line:
            continue
        cells = _viewer_trim_row([cell.strip() for cell in line.split("|")])
        if len(cells) >= 2:
            rows.append(cells)
    return rows


def _viewer_render_table(rows: list[list[str]], *, first_row_header: bool = True) -> str:
    normalized = [_viewer_trim_row([str(cell or "") for cell in row]) for row in rows]
    normalized = [row for row in normalized if row]
    if not normalized:
        return ""

    column_count = max(len(row) for row in normalized)
    padded = [row + [""] * (column_count - len(row)) for row in normalized]
    header_row = padded[0] if first_row_header else None
    body_rows = padded[1:] if first_row_header and len(padded) > 1 else (padded if not first_row_header else [])

    header_html = ""
    if header_row is not None:
        header_html = (
            "<thead><tr>"
            + "".join(f"<th>{_viewer_multiline_html(cell)}</th>" for cell in header_row)
            + "</tr></thead>"
        )

    if not body_rows and header_row is not None:
        body_rows = [header_row]
        header_html = ""

    body_html = (
        "<tbody>"
        + "".join(
            "<tr>" + "".join(f"<td>{_viewer_multiline_html(cell)}</td>" for cell in row) + "</tr>"
            for row in body_rows
        )
        + "</tbody>"
    )

    return f'<div class="viewer-table-wrap"><table class="viewer-html-table">{header_html}{body_html}</table></div>'


def _viewer_render_docx_html(filepath: str) -> str:
    from docx import Document
    from docx.table import Table
    from docx.text.paragraph import Paragraph

    def iter_block_items(document: Any):
        for child in document.element.body.iterchildren():
            tag = str(child.tag)
            if tag.endswith("}p"):
                yield Paragraph(child, document)
            elif tag.endswith("}tbl"):
                yield Table(child, document)

    doc = Document(filepath)
    parts: list[str] = []
    for block in iter_block_items(doc):
        if isinstance(block, Paragraph):
            text = block.text.strip()
            if not text:
                continue
            style_name = (getattr(block.style, "name", "") or "").lower()
            if style_name in {"title", "subtitle"}:
                tag = "h2"
            elif style_name.startswith("heading"):
                tag = "h3"
            else:
                tag = "p"
            parts.append(f"<{tag}>{_viewer_multiline_html(text)}</{tag}>")
            continue

        rows = []
        for row in block.rows:
            values = _viewer_trim_row([cell.text.strip() for cell in row.cells])
            if values:
                rows.append(values)
        table_html = _viewer_render_table(rows)
        if table_html:
            parts.append(table_html)

    if not parts:
        parts.append('<div class="viewer-inline-empty">표시할 내용이 없습니다.</div>')
    return "".join(parts)


def _viewer_excel_string(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value)


def _viewer_render_excel_html(filepath: str, file_type: str) -> str:
    max_rows = 300
    max_cols = 32
    sections: list[str] = []
    empty_html = '<div class="viewer-inline-empty">표시할 데이터가 없습니다.</div>'

    if file_type == "xlsx":
        import openpyxl

        workbook = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        try:
            sheet_names = list(workbook.sheetnames)
            for sheet_index, sheet_name in enumerate(sheet_names, start=1):
                worksheet = workbook[sheet_name]
                rows: list[list[str]] = []
                truncated = False
                for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                    values = _viewer_trim_row([_viewer_excel_string(value).strip() for value in list(row)[:max_cols]])
                    if values:
                        rows.append(values)
                    if row_index >= max_rows:
                        truncated = True
                        break
                table_html = _viewer_render_table(rows)
                note = (
                    '<div class="viewer-inline-note">행 수가 많아 상위 300행까지만 표시합니다.</div>'
                    if truncated
                    else ""
                )
                sections.append(
                    f'<section class="viewer-html-page">'
                    f'<div class="viewer-page-meta">시트 {sheet_index} · {_viewer_html(sheet_name)}</div>'
                    f'<div class="viewer-html-content">{table_html or empty_html}{note}</div>'
                    f"</section>"
                )
        finally:
            workbook.close()
        return "".join(sections)

    import xlrd

    workbook = xlrd.open_workbook(filepath, on_demand=True)
    try:
        for sheet_index in range(workbook.nsheets):
            worksheet = workbook.sheet_by_index(sheet_index)
            rows = []
            truncated = False
            for row_index in range(min(worksheet.nrows, max_rows)):
                values = _viewer_trim_row([
                    _viewer_excel_string(worksheet.cell_value(row_index, col_index)).strip()
                    for col_index in range(min(worksheet.ncols, max_cols))
                ])
                if values:
                    rows.append(values)
            if worksheet.nrows > max_rows:
                truncated = True
            table_html = _viewer_render_table(rows)
            note = (
                '<div class="viewer-inline-note">행 수가 많아 상위 300행까지만 표시합니다.</div>'
                if truncated
                else ""
            )
            sections.append(
                f'<section class="viewer-html-page">'
                f'<div class="viewer-page-meta">시트 {sheet_index + 1} · {_viewer_html(worksheet.name)}</div>'
                f'<div class="viewer-html-content">{table_html or empty_html}{note}</div>'
                f"</section>"
            )
    finally:
        workbook.release_resources()

    return "".join(sections)


def _viewer_render_transformed_html(parsed: dict[str, Any]) -> str:
    pages = parsed.get("pages") or []
    sections: list[str] = []

    for index, page in enumerate(pages, start=1):
        page_num = int(page.get("page_num") or index)
        page_label = page.get("sheet_name") or f"페이지 {page_num}"
        blocks = page.get("blocks") or []
        body_parts: list[str] = []

        for block in blocks:
            text = str(block.get("text") or "").strip()
            if not text:
                continue
            block_type = str(block.get("type") or "").lower()
            if block_type == "title":
                body_parts.append(f'<h3 class="viewer-block-title">{_viewer_multiline_html(text)}</h3>')
            elif block_type == "table":
                rows = _viewer_table_rows_from_pipe_text(text)
                table_html = _viewer_render_table(rows) if rows else ""
                body_parts.append(table_html or f'<pre class="viewer-pre">{_viewer_html(text)}</pre>')
            else:
                body_parts.append(f"<p>{_viewer_multiline_html(text)}</p>")

        if not body_parts:
            fallback_text = str(page.get("text") or page.get("rag_text") or "").strip()
            if fallback_text:
                body_parts.append(f'<pre class="viewer-pre">{_viewer_html(fallback_text)}</pre>')
            else:
                body_parts.append('<div class="viewer-inline-empty">표시할 내용이 없습니다.</div>')

        sections.append(
            f'<section class="viewer-html-page">'
            f'<div class="viewer-page-meta">{_viewer_html(page_label)}</div>'
            f'<div class="viewer-html-content">{"".join(body_parts)}</div>'
            f"</section>"
        )

    if not sections:
        sections.append('<div class="viewer-inline-empty">표시할 내용이 없습니다.</div>')
    return "".join(sections)


def _viewer_payload_v2(parsed: dict[str, Any]) -> dict[str, Any]:
    doc_id = str(parsed.get("id") or "")
    file_type = str(parsed.get("file_type") or "").lower()
    filename = str(parsed.get("filename") or "")
    filepath = str(parsed.get("filepath") or "")
    pages = parsed.get("pages") or []

    if file_type == "pdf":
        page_entries = []
        for index, page in enumerate(pages, start=1):
            page_num = int(page.get("page_num") or index)
            page_entries.append(
                {
                    "page_num": page_num,
                    "preview_url": f"/api/documents/{doc_id}/pages/{page_num}/preview",
                }
            )
        return {
            "mode": "pdf_previews",
            "title": filename,
            "subtitle": f"PDF 미리보기 · {len(page_entries)} pages",
            "file_type": file_type,
            "pages": page_entries,
        }

    html = ""
    subtitle = "읽기용 변환 문서"
    if file_type == "docx":
        try:
            html = _viewer_render_docx_html(filepath)
            subtitle = "DOCX 라이브러리 렌더링"
        except Exception as exc:
            logger.warning("DOCX viewer fallback for %s: %s", filename, exc)
            html = _viewer_render_transformed_html(parsed)
            subtitle = "DOCX 변환 뷰"
    elif file_type in {"xlsx", "xls"}:
        try:
            html = _viewer_render_excel_html(filepath, file_type)
            subtitle = f"{file_type.upper()} 라이브러리 렌더링"
        except Exception as exc:
            logger.warning("Excel viewer fallback for %s: %s", filename, exc)
            html = _viewer_render_transformed_html(parsed)
            subtitle = f"{file_type.upper()} 변환 뷰"
    elif file_type in {"doc", "hwp"}:
        html = _viewer_render_transformed_html(parsed)
        subtitle = f"{file_type.upper()} 변환 뷰"
    else:
        html = _viewer_render_transformed_html(parsed)

    return {
        "mode": "html_document",
        "title": filename,
        "subtitle": subtitle,
        "file_type": file_type,
        "html": html,
    }


def _viewer_payload(parsed: dict[str, Any]) -> dict[str, Any]:
    doc_id = str(parsed.get("id") or "")
    file_type = str(parsed.get("file_type") or "").lower()
    filename = str(parsed.get("filename") or "")
    filepath = str(parsed.get("filepath") or "")
    pages = parsed.get("pages") or []

    if file_type == "pdf":
        page_entries = []
        for index, page in enumerate(pages, start=1):
            page_num = int(page.get("page_num") or index)
            page_entries.append(
                {
                    "page_num": page_num,
                    "preview_url": f"/api/documents/{doc_id}/pages/{page_num}/preview",
                }
            )
        return {
            "mode": "pdf_previews",
            "title": filename,
            "subtitle": f"PDF 미리보기 · {len(page_entries)} pages",
            "file_type": file_type,
            "pages": page_entries,
        }

    html = ""
    subtitle = "변환 문서 뷰"
    render_source = "converted"
    if file_type == "docx":
        try:
            html = _viewer_render_docx_html(filepath)
            subtitle = "DOCX 라이브러리 렌더"
            render_source = "library"
        except Exception as exc:
            logger.warning("DOCX viewer fallback for %s: %s", filename, exc)
            html = _viewer_render_transformed_html(parsed)
            subtitle = "DOCX 변환 뷰 (fallback)"
            render_source = "converted_fallback"
    elif file_type in {"xlsx", "xls"}:
        try:
            html = _viewer_render_excel_html(filepath, file_type)
            subtitle = f"{file_type.upper()} 라이브러리 렌더"
            render_source = "library"
        except Exception as exc:
            logger.warning("Excel viewer fallback for %s: %s", filename, exc)
            html = _viewer_render_transformed_html(parsed)
            subtitle = f"{file_type.upper()} 변환 뷰 (fallback)"
            render_source = "converted_fallback"
    elif file_type in {"doc", "hwp"}:
        html = _viewer_render_transformed_html(parsed)
        subtitle = f"{file_type.upper()} 변환 뷰"
        render_source = "converted"
    else:
        html = _viewer_render_transformed_html(parsed)

    return {
        "mode": "html_document",
        "title": filename,
        "subtitle": subtitle,
        "file_type": file_type,
        "render_source": render_source,
        "html": html,
    }


def _viewer_payload_with_sidecar(parsed: dict[str, Any]) -> dict[str, Any]:
    doc_id = _parsed_doc_id(parsed)
    file_type = str(parsed.get("file_type") or "").lower()
    filename = str(parsed.get("filename") or "")
    filepath = str(parsed.get("filepath") or "")
    pages = parsed.get("pages") or []

    pdf_source = _viewer_pdf_source_path(
        parsed,
        generate_if_missing=_viewer_sidecar_generation_allowed(file_type),
    )
    if pdf_source is not None and pdf_source.is_file():
        page_count = _pdf_page_count(pdf_source) or len(pages) or 1
        page_entries = [
            {
                "page_num": page_num,
                "preview_url": f"/api/documents/{doc_id}/pages/{page_num}/preview",
            }
            for page_num in range(1, page_count + 1)
        ]
        source_label = "PDF" if file_type == "pdf" else "PDF sidecar"
        return {
            "mode": "pdf_previews",
            "title": filename,
            "subtitle": f"{source_label} preview - {len(page_entries)} pages",
            "file_type": file_type,
            "pages": page_entries,
            "stitched_preview_url": (
                f"/api/documents/{doc_id}/preview-strip"
                if file_type in {"xls", "xlsx"} and page_count > 1
                else ""
            ),
        }

    html = ""
    subtitle = "Converted document view"
    render_source = "converted"
    if file_type == "docx":
        try:
            html = _viewer_render_docx_html(filepath)
            subtitle = "DOCX library render"
            render_source = "library"
        except Exception as exc:
            logger.warning("DOCX viewer fallback for %s: %s", filename, exc)
            html = _viewer_render_transformed_html(parsed)
            subtitle = "DOCX converted view (fallback)"
            render_source = "converted_fallback"
    elif file_type in {"xlsx", "xls"}:
        try:
            html = _viewer_render_excel_html(filepath, file_type)
            subtitle = f"{file_type.upper()} library render"
            render_source = "library"
        except Exception as exc:
            logger.warning("Excel viewer fallback for %s: %s", filename, exc)
            html = _viewer_render_transformed_html(parsed)
            subtitle = f"{file_type.upper()} converted view (fallback)"
            render_source = "converted_fallback"
    elif file_type in {"doc", "hwp"}:
        html = _viewer_render_transformed_html(parsed)
        subtitle = f"{file_type.upper()} converted view"
        render_source = "converted"
    else:
        html = _viewer_render_transformed_html(parsed)

    return {
        "mode": "html_document",
        "title": filename,
        "subtitle": subtitle,
        "file_type": file_type,
        "render_source": render_source,
        "html": html,
    }


def _save_result(doc_id: str, data: dict[str, Any]) -> None:
    try:
        save_all_parse_outputs(doc_id, data, RESULT_DIR, PARSER_VERSION)
    except Exception as exc:
        logger.warning("Failed to persist result %s: %s", doc_id, exc)


def _is_legacy_result(data: dict[str, Any]) -> bool:
    """Check if the cached JSON is from the older version of the parser."""
    if data.get("parser_version") != PARSER_VERSION:
        return True

    pages = data.get("pages", [])
    if not pages:
        return False

    first_page = pages[0]
    file_type = str(data.get("file_type") or "").lower()

    if file_type == "pdf":
        required_fields = ("blocks", "preview_image", "parser_debug", "page_width")
    elif file_type in {"doc", "docx", "hwp"}:
        required_fields = ("blocks", "tables", "text")
    elif file_type in {"xlsx", "xls"}:
        required_fields = ("sheet_name", "text", "tables", "row_count", "col_count")
    else:
        required_fields = ("text",)

    return not all(field in first_page for field in required_fields)


def _load_result(doc_id: str) -> dict[str, Any] | None:
    p = _result_path(doc_id)
    if not p.exists():
        return None
    try:
        with open(p, "r", encoding="utf-8") as f:
            data = json.load(f)
            if _is_legacy_result(data):
                logger.info("Ignoring legacy cached result for %s", doc_id)
                return None
            return data
    except Exception as exc:
        logger.warning("Failed to read cached result %s: %s", doc_id, exc)
        return None


def _warm_cache() -> None:
    """Load every .json file in RESULT_DIR into the memory cache."""
    for name in os.listdir(RESULT_DIR):
        if not name.endswith(".json"):
            continue
        did = name[:-5]
        if did in parsed_cache:
            continue
        data = _load_result(did)
        if data:
            parsed_cache[did] = data
    logger.info("Loaded %d cached results from disk", len(parsed_cache))


def _scan_documents() -> list[dict[str, Any]]:
    docs: list[dict[str, Any]] = []
    if not os.path.isdir(DOC_DIR):
        return docs
    for fname in sorted(os.listdir(DOC_DIR)):
        ext = os.path.splitext(fname)[1].lower()
        if ext not in SUPPORTED_EXTENSIONS:
            continue
        fpath = os.path.join(DOC_DIR, fname)
        if not os.path.isfile(fpath):
            continue
        did = _doc_id(fpath)
        cached = parsed_cache.get(did)
        docs.append({
            "id": did,
            "filename": fname,
            "file_type": ext.lstrip("."),
            "file_size": os.path.getsize(fpath),
            "parsed": cached is not None,
            "status": cached["status"] if cached else "pending",
            "quality_score": cached.get("metadata", {}).get("quality_score") if cached else None,
            "quality_grade": cached.get("metadata", {}).get("quality_grade") if cached else None,
        })
    return docs


def _read_json_file(path: Path) -> dict[str, Any]:
    try:
        with path.open("r", encoding="utf-8") as handle:
            payload = json.load(handle)
    except FileNotFoundError as exc:
        raise HTTPException(404, f"Required file not found: {path}") from exc
    except Exception as exc:
        raise HTTPException(500, f"Could not read JSON file: {path}") from exc
    if not isinstance(payload, dict):
        raise HTTPException(500, f"Expected JSON object: {path}")
    return payload


def _resolve_project_path(
    raw_path: Any,
    artifact_root_name: str = "chroma_indexes",
    run_id: Any = None,
    fallback_filename: str | None = None,
) -> Path:
    if raw_path:
        path = Path(str(raw_path))
    elif run_id and fallback_filename:
        return (PROJECT_ROOT / artifact_root_name / str(run_id) / fallback_filename).resolve()
    else:
        raise HTTPException(500, "Missing artifact path")

    windows_path = PureWindowsPath(str(raw_path)) if raw_path else None
    path_parts = windows_path.parts if windows_path and windows_path.is_absolute() else path.parts

    if not path.is_absolute() and not (windows_path and windows_path.is_absolute()):
        if path.parts and path.parts[0] == PROJECT_ROOT.name:
            return (PROJECT_ROOT.parent / path).resolve()
        return (PROJECT_ROOT / path).resolve()

    if path.exists():
        return path

    artifact_root = PROJECT_ROOT / artifact_root_name
    if artifact_root_name in path_parts:
        tail = Path(*path_parts[path_parts.index(artifact_root_name) + 1 :])
        candidate = artifact_root / tail
        if candidate.exists():
            return candidate

    fallback_run_id = str(run_id) if run_id else next((part for part in reversed(path_parts) if part.startswith("run_")), None)
    if fallback_run_id:
        if fallback_run_id in path_parts:
            tail_parts = path_parts[path_parts.index(fallback_run_id) + 1 :]
            tail = Path(*tail_parts) if tail_parts else Path(fallback_filename or path.name)
        else:
            tail = Path(fallback_filename or path.name)
        candidate = artifact_root / fallback_run_id / tail
        if candidate.exists():
            return candidate
        if fallback_filename:
            candidate = artifact_root / fallback_run_id / fallback_filename
            if candidate.exists():
                return candidate

    return path


def _rag_api_post(endpoint: str, payload: dict[str, Any]) -> dict[str, Any]:
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    http_request = request.Request(
        f"{RAG_API_BASE_URL}{endpoint}",
        data=body,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        with request.urlopen(http_request, timeout=180) as response:
            result = json.loads(response.read().decode("utf-8"))
    except error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise HTTPException(exc.code, detail) from exc
    except error.URLError as exc:
        raise HTTPException(502, f"RAG API is not reachable at {RAG_API_BASE_URL}: {exc.reason}") from exc
    except Exception as exc:
        raise HTTPException(500, f"RAG API proxy failed: {exc}") from exc
    if not isinstance(result, dict):
        raise HTTPException(502, "RAG API returned a non-object response")
    return result


def _rag_api_stream(
    endpoint: str,
    payload: dict[str, Any],
    cache_key: tuple[Any, ...] | None = None,
):
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    http_request = request.Request(
        f"{RAG_API_BASE_URL}{endpoint}",
        data=body,
        headers={
            "Content-Type": "application/json",
            "Accept": "text/event-stream",
            "Cache-Control": "no-cache",
        },
        method="POST",
    )

    def _stream():
        try:
            with request.urlopen(http_request, timeout=180) as response:
                current_event = "message"
                data_lines: list[str] = []
                for raw_line in response:
                    yield raw_line
                    line = raw_line.decode("utf-8", errors="replace")
                    if line.startswith("event:"):
                        current_event = line.split(":", 1)[1].strip() or "message"
                    elif line.startswith("data:"):
                        data_lines.append(line.split(":", 1)[1].lstrip())
                    elif not line.strip():
                        if current_event == "done" and cache_key and endpoint == "/qa/stream" and data_lines:
                            payload_text = "".join(data_lines).strip()
                            try:
                                cached_payload = json.loads(payload_text)
                            except json.JSONDecodeError:
                                cached_payload = None
                            if isinstance(cached_payload, dict):
                                with rag_response_cache_lock:
                                    qa_response_cache[cache_key] = cached_payload
                        current_event = "message"
                        data_lines = []
        except error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")
            yield _sse_message("error", {"detail": detail or f"Upstream returned HTTP {exc.code}."})
        except error.URLError as exc:
            yield _sse_message("error", {"detail": f"RAG API is not reachable at {RAG_API_BASE_URL}: {exc.reason}"})
        except Exception as exc:
            yield _sse_message("error", {"detail": f"RAG API proxy failed: {exc}"})

    return _stream()


def _sse_message(event: str, payload: dict[str, Any]) -> bytes:
    body = json.dumps(payload, ensure_ascii=False)
    return f"event: {event}\ndata: {body}\n\n".encode("utf-8")


def _stream_cached_qa_response(payload: dict[str, Any]):
    def _stream():
        yield _sse_message("retrieved", {"count": len(payload.get("sources") or []), "message": "Loaded cached QA response."})
        yield _sse_message("done", payload)

    return _stream()


def _is_document_title_query(query: str) -> bool:
    normalized = "".join(str(query or "").lower().split())
    title_terms = ("문서제목", "문서명", "파일명", "제목이뭐", "제목뭐", "documenttitle", "filename")
    return any(term in normalized for term in title_terms)


def _document_title_response(query: str, filename: str) -> dict[str, Any]:
    return {
        "mode": "qa",
        "title": query,
        "answer": f"선택된 문서 제목은 `{filename}`입니다.",
        "sources": [{"document": filename, "page": 0}],
    }


def _load_rag_documents() -> list[dict[str, Any]]:
    latest_path = PROJECT_ROOT / "chroma_indexes" / "latest.json"
    id_manifest: list[dict[str, Any]] = []
    try:
        latest = _read_json_file(latest_path)
    except HTTPException as exc:
        logger.warning("Could not load RAG latest index %s: %s", latest_path, exc.detail)
        latest = None

    if latest is not None:
        id_manifest_path = _resolve_project_path(
            latest.get("id_manifest_path"),
            artifact_root_name="chroma_indexes",
            run_id=latest.get("run_id"),
            fallback_filename="id_manifest.json",
        )
        try:
            with id_manifest_path.open("r", encoding="utf-8") as handle:
                loaded_manifest = json.load(handle)
        except FileNotFoundError:
            logger.warning("Could not find RAG id manifest: %s", id_manifest_path)
        except Exception as exc:
            logger.warning("Could not read RAG id manifest %s: %s", id_manifest_path, exc)
        else:
            if isinstance(loaded_manifest, list):
                id_manifest = loaded_manifest
            else:
                logger.warning("Expected JSON array in RAG id manifest: %s", id_manifest_path)

    grouped: dict[str, dict[str, Any]] = {}
    page_counts: dict[str, set[int]] = {}
    strategy_counts: dict[str, Counter[str]] = {}
    for item in id_manifest:
        if not isinstance(item, dict):
            continue
        filename = str(item.get("filename") or "").strip()
        if not filename:
            continue
        doc = grouped.setdefault(
            filename,
            {
                "filename": filename,
                "doc_id": str(item.get("doc_id") or ""),
                "chunk_count": 0,
                "page_count": 0,
                "strategies": [],
            },
        )
        doc["chunk_count"] += 1
        try:
            page_num = int(item.get("page_num") or 0)
        except (TypeError, ValueError):
            page_num = 0
        page_counts.setdefault(filename, set()).add(page_num)
        strategy_counts.setdefault(filename, Counter())[str(item.get("strategy_name") or "unknown")] += 1

    docs: list[dict[str, Any]] = []
    for filename, doc in grouped.items():
        pages = {page for page in page_counts.get(filename, set()) if page > 0}
        strategies = sorted(strategy_counts.get(filename, Counter()).keys())
        docs.append(
            {
                **doc,
                "page_count": len(pages),
                "strategies": strategies,
                "status": "ready",
                "stage": "done",
                "progress": 100,
                "message": "사용 가능",
                "is_pending": False,
                "is_available": True,
            }
        )

    docs_by_filename = {str(doc.get("filename") or ""): doc for doc in docs}
    with upload_jobs_lock:
        pending_upload_jobs = [
            dict(job)
            for job in upload_jobs.values()
            if str(job.get("kind") or "") == "upload"
            and str(job.get("status") or "") in {"queued", "running"}
        ]

    for job in pending_upload_jobs:
        filename = str(job.get("filename") or "").strip()
        if not filename:
            continue

        doc_id = str(job.get("doc_id") or "")
        parsed = None
        if doc_id:
            parsed = parsed_cache.get(doc_id)
            if parsed is None:
                parsed = _load_result(doc_id)
                if parsed is not None:
                    parsed_cache[doc_id] = parsed

        base = dict(docs_by_filename.get(filename) or {})
        page_count = len(parsed.get("pages") or []) if parsed else int(base.get("page_count") or 0)
        file_type = str((parsed or {}).get("file_type") or Path(filename).suffix.lstrip(".").lower())
        base.update(
            {
                "filename": filename,
                "doc_id": doc_id or str(base.get("doc_id") or ""),
                "file_type": file_type,
                "chunk_count": int(base.get("chunk_count") or 0),
                "page_count": page_count,
                "strategies": list(base.get("strategies") or []),
                "status": str(job.get("status") or "queued"),
                "stage": str(job.get("stage") or "queued"),
                "progress": int(job.get("progress") or 0),
                "message": str(job.get("message") or "업로드 중..."),
                "is_pending": True,
                "is_available": False,
                "created_at": str(job.get("created_at") or ""),
                "updated_at": str(job.get("updated_at") or ""),
            }
        )
        docs_by_filename[filename] = base

    return sorted(
        docs_by_filename.values(),
        key=lambda item: (not bool(item.get("is_pending")), str(item.get("filename") or "").casefold()),
    )


def _job_snapshot(job_id: str) -> dict[str, Any]:
    with upload_jobs_lock:
        job = upload_jobs.get(job_id)
        if not job:
            raise HTTPException(404, "Upload job not found")
        return dict(job)


def _set_upload_job(job_id: str, **updates: Any) -> None:
    with upload_jobs_lock:
        job = upload_jobs.get(job_id)
        if not job:
            return
        job.update(updates)
        job["updated_at"] = datetime.now().isoformat(timespec="seconds")


def _safe_upload_name(filename: str) -> str:
    name = Path(filename).name.strip()
    if not name:
        raise HTTPException(400, "Uploaded file must have a filename.")
    ext = Path(name).suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        allowed = ", ".join(sorted(SUPPORTED_EXTENSIONS))
        raise HTTPException(400, f"Unsupported file extension: {ext}. Allowed: {allowed}")
    return name


def _unique_upload_path(filename: str) -> Path:
    documents_dir = Path(DOC_DIR)
    documents_dir.mkdir(parents=True, exist_ok=True)
    candidate = documents_dir / filename
    if not candidate.exists():
        return candidate

    stem = candidate.stem
    suffix = candidate.suffix
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    for index in range(1, 1000):
        renamed = documents_dir / f"{stem}_{stamp}_{index}{suffix}"
        if not renamed.exists():
            return renamed
    raise HTTPException(500, "Could not allocate a unique upload filename.")


def _run_pipeline_command(job_id: str, stage: str, progress: int, message: str, args: list[str]) -> None:
    _set_upload_job(job_id, stage=stage, progress=progress, message=message)
    logger.info("Upload job %s running: %s", job_id, " ".join(args))
    completed = subprocess.run(
        args,
        cwd=PROJECT_ROOT,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    if completed.returncode != 0:
        detail = (completed.stderr or completed.stdout or "").strip()
        raise RuntimeError(f"{message} failed with exit code {completed.returncode}: {detail}")
    if completed.stdout:
        logger.info("Upload job %s output: %s", job_id, completed.stdout.strip())


def _json_dump_file(path: Path, payload: dict[str, Any] | list[Any]) -> None:
    temp_path = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    with temp_path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)
        handle.write("\n")
    temp_path.replace(path)


def _write_jsonl_file(path: Path, records: list[dict[str, Any]]) -> None:
    temp_path = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    with temp_path.open("w", encoding="utf-8") as handle:
        for record in records:
            handle.write(json.dumps(record, ensure_ascii=False, separators=(",", ":")))
            handle.write("\n")
    temp_path.replace(path)


def _read_jsonl_file(path: Path) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as handle:
        for line_number, line in enumerate(handle, start=1):
            stripped = line.strip()
            if not stripped:
                continue
            payload = json.loads(stripped)
            if not isinstance(payload, dict):
                raise RuntimeError(f"Expected JSON object at {path}:{line_number}")
            records.append(payload)
    return records


def _portable_project_path(path: Path) -> str:
    try:
        return path.resolve().relative_to(PROJECT_ROOT.resolve()).as_posix()
    except ValueError:
        return str(path)


def _make_run_dir(output_dir: Path) -> tuple[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    for _ in range(1000):
        run_id = f"run_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}"
        run_dir = output_dir / run_id
        try:
            run_dir.mkdir(parents=False, exist_ok=False)
            return run_id, run_dir
        except FileExistsError:
            continue
    raise RuntimeError(f"Could not allocate run directory under {output_dir}")


def _create_filtered_vector_run_excluding_doc(job_id: str, doc_id: str) -> dict[str, Any]:
    if not doc_id:
        raise RuntimeError("Cannot rebuild vector index without doc_id.")

    import numpy as np

    vector_indexes_dir = PROJECT_ROOT / "vector_indexes"
    latest_path = vector_indexes_dir / "latest.json"
    latest = _read_json_file(latest_path)
    source_run_id = str(latest.get("run_id") or "")
    manifest_path = _resolve_project_path(
        latest.get("manifest_path"),
        artifact_root_name="vector_indexes",
        run_id=source_run_id,
        fallback_filename="run_manifest.json",
    )
    metadata_path = _resolve_project_path(
        latest.get("metadata_path"),
        artifact_root_name="vector_indexes",
        run_id=source_run_id,
        fallback_filename="chunk_metadata_records.jsonl",
    )
    vectors_path = _resolve_project_path(
        latest.get("vectors_path"),
        artifact_root_name="vector_indexes",
        run_id=source_run_id,
        fallback_filename="vectors.npz",
    )
    id_map_path = _resolve_project_path(
        latest.get("id_map_path"),
        artifact_root_name="vector_indexes",
        run_id=source_run_id,
        fallback_filename="id_map.json",
    )

    missing_paths = [path for path in (manifest_path, metadata_path, vectors_path, id_map_path) if not path.exists()]
    if missing_paths:
        raise RuntimeError("Missing vector index artifact(s): " + ", ".join(str(path) for path in missing_paths))

    source_manifest = _read_json_file(manifest_path)
    metadata_records = _read_jsonl_file(metadata_path)
    source_id_map = _read_json_file(id_map_path)
    vector_data = np.load(vectors_path, allow_pickle=True)
    vectors = vector_data["vectors"]
    vector_ids = [str(item) for item in vector_data["vector_record_ids"].tolist()]
    if len(vector_ids) != len(vectors):
        raise RuntimeError(f"Vector id count mismatch: {len(vector_ids)} ids for {len(vectors)} vectors.")

    records_by_vector_id = {str(record.get("vector_record_id")): record for record in metadata_records}
    keep_indices: list[int] = []
    filtered_vector_ids: list[str] = []
    for index, vector_id in enumerate(vector_ids):
        record = records_by_vector_id.get(vector_id)
        mapped = source_id_map.get(str(index))
        mapped_doc_id = str(mapped.get("doc_id") or "") if isinstance(mapped, dict) else ""
        record_doc_id = str(record.get("doc_id") or "") if record else mapped_doc_id
        if record_doc_id == doc_id:
            continue
        keep_indices.append(index)
        filtered_vector_ids.append(vector_id)

    filtered_metadata = [record for record in metadata_records if str(record.get("doc_id") or "") != doc_id]
    if len(keep_indices) == len(vector_ids):
        raise RuntimeError(f"Document {doc_id} was not present in the latest vector index.")

    if keep_indices:
        filtered_vectors = vectors[keep_indices]
    else:
        dimensions = int(vectors.shape[1]) if getattr(vectors, "ndim", 0) == 2 else 0
        filtered_vectors = np.empty((0, dimensions), dtype=vectors.dtype)

    filtered_id_map: dict[str, dict[str, str]] = {}
    for new_index, vector_id in enumerate(filtered_vector_ids):
        record = records_by_vector_id.get(vector_id)
        if record is None:
            old_index = keep_indices[new_index]
            mapped = source_id_map.get(str(old_index))
            record = mapped if isinstance(mapped, dict) else {}
        filtered_id_map[str(new_index)] = {
            "vector_record_id": str(vector_id),
            "chunk_id": str(record.get("chunk_id") or ""),
            "doc_id": str(record.get("doc_id") or ""),
            "strategy_name": str(record.get("strategy_name") or ""),
        }

    run_id, run_dir = _make_run_dir(vector_indexes_dir)
    input_chunk_files = [
        str(path)
        for path in source_manifest.get("input_chunk_files", [])
        if doc_id not in str(path)
    ]
    strategy_distribution = Counter(str(record.get("strategy_name") or "unknown") for record in filtered_metadata)
    doc_distribution = Counter(str(record.get("doc_id") or "unknown") for record in filtered_metadata)
    embedding_dimensions = int(filtered_vectors.shape[1]) if filtered_vectors.ndim == 2 else 0
    manifest = {
        "run_id": run_id,
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "gpt_model": source_manifest.get("gpt_model", ""),
        "embedding_model": source_manifest.get("embedding_model", ""),
        "embedding_dimensions": embedding_dimensions,
        "input_chunk_files": input_chunk_files,
        "total_input_chunks": len(filtered_metadata),
        "embedded_chunks": len(filtered_vector_ids),
        "skipped_empty_chunks": max(0, len(filtered_metadata) - len(filtered_vector_ids)),
        "strategy_distribution": dict(strategy_distribution),
        "doc_distribution": dict(doc_distribution),
        "operation": "delete_doc_filter",
        "source_vector_run_id": source_run_id,
        "excluded_doc_id": doc_id,
        "job_id": job_id,
    }
    embedding_stats = {
        **manifest,
        "source_metadata_count": len(metadata_records),
        "source_vector_count": len(vector_ids),
        "removed_metadata_count": len(metadata_records) - len(filtered_metadata),
        "removed_vector_count": len(vector_ids) - len(filtered_vector_ids),
        "embedding_batch_size": 0,
        "skipped_empty_vector_record_ids": [],
    }

    _json_dump_file(run_dir / "run_manifest.json", manifest)
    _write_jsonl_file(run_dir / "chunk_metadata_records.jsonl", filtered_metadata)
    _json_dump_file(run_dir / "embedding_stats.json", embedding_stats)
    np.savez_compressed(
        run_dir / "vectors.npz",
        vectors=filtered_vectors,
        vector_record_ids=np.asarray(filtered_vector_ids, dtype=object),
    )
    _json_dump_file(run_dir / "id_map.json", filtered_id_map)
    latest_payload = {
        "run_id": run_id,
        "run_dir": _portable_project_path(run_dir),
        "manifest_path": _portable_project_path(run_dir / "run_manifest.json"),
        "metadata_path": _portable_project_path(run_dir / "chunk_metadata_records.jsonl"),
        "vectors_path": _portable_project_path(run_dir / "vectors.npz"),
        "id_map_path": _portable_project_path(run_dir / "id_map.json"),
    }
    logger.info(
        "Delete job %s created filtered vector run %s, removed %s vectors for doc_id=%s",
        job_id,
        run_id,
        len(vector_ids) - len(filtered_vector_ids),
        doc_id,
    )
    return {"run_id": run_id, "latest_payload": latest_payload, "manifest": manifest}


def _process_upload_job(job_id: str, file_path: Path) -> None:
    try:
        with index_build_lock:
            _set_upload_job(job_id, status="running", stage="parse", progress=10, message="문서 파싱 중...")
            doc_id = _doc_id(str(file_path))
            result = parse_document(str(file_path))
            if result.get("status") == "error":
                raise RuntimeError(str(result.get("error") or "Parser returned an error."))
            _save_result(doc_id, result)
            try:
                write_root_llm_exports(result, PROJECT_ROOT)
            except Exception as exc:
                logger.warning("write_root_llm_exports failed for upload job %s: %s", job_id, exc)
            parsed_cache[doc_id] = result
            _maybe_create_sidecar_pdf(result)

            _run_pipeline_command(
                job_id,
                stage="chunk",
                progress=35,
                message="청크 생성 중...",
                args=[sys.executable, "-m", "scripts.build_chunks", "--doc-id", doc_id],
            )
            _run_pipeline_command(
                job_id,
                stage="embedding",
                progress=60,
                message="임베딩 생성 중...",
                args=[sys.executable, "-m", "scripts.run_retrieval_prep"],
            )
            _run_pipeline_command(
                job_id,
                stage="chroma",
                progress=85,
                message="Chroma 인덱스 반영 중...",
                args=[sys.executable, "-m", "scripts.run_chroma_ingest"],
            )

        _set_upload_job(
            job_id,
            status="completed",
            stage="done",
            progress=100,
            message="업로드와 인덱싱이 완료되었습니다.",
            doc_id=doc_id,
            filename=file_path.name,
        )
        _invalidate_rag_response_cache(filename=file_path.name)
    except Exception as exc:
        logger.exception("Upload job %s failed", job_id)
        _set_upload_job(
            job_id,
            status="failed",
            stage="failed",
            progress=100,
            message="업로드 처리에 실패했습니다.",
            error=str(exc),
        )


def _managed_delete_roots() -> tuple[Path, ...]:
    return (Path(DOC_DIR), Path(RESULT_DIR), PROJECT_ROOT / "chunks")


def _resolve_managed_path(path: Path, roots: tuple[Path, ...]) -> tuple[Path, Path]:
    resolved = path.resolve()
    allowed_roots = tuple(root.resolve() for root in roots)
    for root in allowed_roots:
        if resolved == root or root in resolved.parents:
            return resolved, root
    raise RuntimeError(f"Refusing to manage path outside allowed directories: {resolved}")


def _stage_path_for_delete(
    path: Path,
    roots: tuple[Path, ...],
    trash_dir: Path,
    staged_paths: list[dict[str, Path]],
) -> None:
    resolved, root = _resolve_managed_path(path, roots)
    if not resolved.exists():
        return
    relative = resolved.relative_to(root)
    if relative == Path("."):
        raise RuntimeError(f"Refusing to stage managed root directory: {resolved}")

    destination = trash_dir / root.name / relative
    if destination.exists():
        destination = destination.with_name(f"{destination.name}.{uuid.uuid4().hex[:8]}")
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.move(str(resolved), str(destination))
    staged_paths.append({"source": resolved, "trash": destination})


def _restore_conflict_path(path: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    suffix = f".restore_conflict_{stamp}_{uuid.uuid4().hex[:8]}"
    if path.suffix:
        candidate = path.with_name(f"{path.stem}{suffix}{path.suffix}")
    else:
        candidate = path.with_name(f"{path.name}{suffix}")
    while candidate.exists():
        candidate = candidate.with_name(f"{candidate.name}.{uuid.uuid4().hex[:4]}")
    return candidate


def _restore_staged_delete(staged_delete: dict[str, Any]) -> None:
    for item in reversed(staged_delete.get("staged_paths", [])):
        trash_path = Path(item["trash"])
        source_path = Path(item["source"])
        if not trash_path.exists():
            continue
        restore_path = source_path if not source_path.exists() else _restore_conflict_path(source_path)
        restore_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.move(str(trash_path), str(restore_path))

    doc_id = str(staged_delete.get("doc_id") or "")
    cache_entry = staged_delete.get("cache_entry")
    if doc_id and cache_entry is not None:
        parsed_cache[doc_id] = cache_entry


def _finalize_staged_delete(staged_delete: dict[str, Any]) -> None:
    trash_dir = Path(staged_delete["trash_dir"])
    trash_root = Path(staged_delete["trash_root"])
    resolved_dir = trash_dir.resolve()
    resolved_root = trash_root.resolve()
    if resolved_dir == resolved_root or resolved_root not in resolved_dir.parents:
        raise RuntimeError(f"Refusing to delete unsafe trash path: {resolved_dir}")
    if resolved_dir.exists():
        shutil.rmtree(resolved_dir)


def _stage_document_artifacts_for_delete(filename: str, doc_id: str, job_id: str) -> dict[str, Any]:
    roots = _managed_delete_roots()
    documents_dir = Path(DOC_DIR)
    parsed_dir = Path(RESULT_DIR)
    chunks_dir = PROJECT_ROOT / "chunks"
    trash_root = PROJECT_ROOT / "rag_api_runs" / "delete_trash"
    trash_dir = trash_root / job_id
    trash_dir.mkdir(parents=True, exist_ok=False)

    staged_paths: list[dict[str, Path]] = []
    cache_entry = parsed_cache.pop(doc_id, None) if doc_id else None
    staged_delete: dict[str, Any] = {
        "doc_id": doc_id,
        "trash_root": trash_root,
        "trash_dir": trash_dir,
        "staged_paths": staged_paths,
        "cache_entry": cache_entry,
    }

    try:
        _stage_path_for_delete(documents_dir / filename, roots, trash_dir, staged_paths)
        if doc_id:
            _stage_path_for_delete(parsed_dir / f"{doc_id}.json", roots, trash_dir, staged_paths)
            review_dir = parsed_dir / "review"
            if review_dir.exists():
                for path in sorted(review_dir.glob(f"{doc_id}*")):
                    _stage_path_for_delete(path, roots, trash_dir, staged_paths)
            preview_dir = parsed_dir / "previews" / doc_id
            _stage_path_for_delete(preview_dir, roots, trash_dir, staged_paths)
            if chunks_dir.exists():
                for path in sorted(chunks_dir.glob(f"{doc_id}.*.json")):
                    _stage_path_for_delete(path, roots, trash_dir, staged_paths)
    except Exception:
        _restore_staged_delete(staged_delete)
        raise

    return staged_delete


def _process_delete_job(job_id: str, filename: str, doc_id: str) -> None:
    staged_delete: dict[str, Any] | None = None
    vector_latest_backup: dict[str, Any] | None = None
    chroma_latest_backup: dict[str, Any] | None = None
    try:
        with index_build_lock:
            _set_upload_job(job_id, status="running", stage="delete", progress=15, message="문서와 기존 산출물을 안전 보관 중...")
            staged_delete = _stage_document_artifacts_for_delete(filename, doc_id, job_id)

            vector_latest_path = PROJECT_ROOT / "vector_indexes" / "latest.json"
            chroma_latest_path = PROJECT_ROOT / "chroma_indexes" / "latest.json"
            if vector_latest_path.exists():
                vector_latest_backup = _read_json_file(vector_latest_path)
            if chroma_latest_path.exists():
                chroma_latest_backup = _read_json_file(chroma_latest_path)

            _set_upload_job(job_id, stage="vector_filter", progress=50, message="기존 임베딩에서 삭제 문서만 제외하는 중...")
            filtered_run = _create_filtered_vector_run_excluding_doc(job_id, doc_id)
            _run_pipeline_command(
                job_id,
                stage="chroma",
                progress=85,
                message="Chroma 인덱스를 삭제 반영본으로 재생성 중...",
                args=[sys.executable, "-m", "scripts.run_chroma_ingest", "--vector-run", str(filtered_run["run_id"])],
            )
            _json_dump_file(vector_latest_path, filtered_run["latest_payload"])
            try:
                _finalize_staged_delete(staged_delete)
            except Exception as cleanup_exc:
                logger.warning("Delete job %s completed but trash cleanup failed: %s", job_id, cleanup_exc)

        _set_upload_job(
            job_id,
            status="completed",
            stage="done",
            progress=100,
            message="문서 삭제와 인덱스 갱신이 완료되었습니다.",
            doc_id=doc_id,
            filename=filename,
        )
        _invalidate_rag_response_cache(filename=filename)
    except Exception as exc:
        logger.exception("Delete job %s failed", job_id)
        if vector_latest_backup is not None:
            try:
                _json_dump_file(PROJECT_ROOT / "vector_indexes" / "latest.json", vector_latest_backup)
            except Exception as restore_exc:
                logger.warning("Could not restore vector latest for delete job %s: %s", job_id, restore_exc)
        if chroma_latest_backup is not None:
            try:
                _json_dump_file(PROJECT_ROOT / "chroma_indexes" / "latest.json", chroma_latest_backup)
            except Exception as restore_exc:
                logger.warning("Could not restore Chroma latest for delete job %s: %s", job_id, restore_exc)
        if staged_delete is not None:
            try:
                _restore_staged_delete(staged_delete)
            except Exception as restore_exc:
                logger.warning("Could not restore staged delete artifacts for job %s: %s", job_id, restore_exc)
        _set_upload_job(
            job_id,
            status="failed",
            stage="failed",
            progress=100,
            message="문서 삭제 처리에 실패했습니다. 가능한 산출물은 원래 위치로 복구했습니다.",
            error=str(exc),
        )


# ── routes ───────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def index():
    return FileResponse(
        os.path.join(os.path.dirname(__file__), "static", "index.html"),
        media_type="text/html",
    )


@app.get("/api/documents")
async def list_documents():
    return _scan_documents()


@app.get("/api/rag/documents")
async def list_rag_documents():
    return _load_rag_documents()


@app.post("/api/rag/upload")
async def upload_rag_document(file: UploadFile = File(...)):
    filename = _safe_upload_name(file.filename or "")
    destination = _unique_upload_path(filename)
    doc_id = _doc_id(str(destination))
    try:
        with destination.open("wb") as handle:
            shutil.copyfileobj(file.file, handle)
    except Exception as exc:
        raise HTTPException(500, f"Could not save uploaded file: {exc}") from exc
    finally:
        await file.close()

    job_id = f"upload_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}"
    with upload_jobs_lock:
        upload_jobs[job_id] = {
            "job_id": job_id,
            "kind": "upload",
            "status": "queued",
            "stage": "queued",
            "message": "업로드 완료. 인덱싱 대기 중...",
            "progress": 5,
            "filename": destination.name,
            "doc_id": doc_id,
            "error": None,
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "updated_at": datetime.now().isoformat(timespec="seconds"),
        }

    worker = threading.Thread(target=_process_upload_job, args=(job_id, destination), daemon=True)
    worker.start()
    return _job_snapshot(job_id)


@app.get("/api/rag/upload-jobs/{job_id}")
async def get_upload_job(job_id: str):
    return _job_snapshot(job_id)


@app.post("/api/rag/delete")
async def delete_rag_document(proxy_request: Request):
    payload = await proxy_request.json()
    if not isinstance(payload, dict):
        raise HTTPException(400, "Expected JSON object")
    filename = str(payload.get("filename") or "").strip()
    if not filename:
        raise HTTPException(400, "filename is required")

    docs = _load_rag_documents()
    match = next((doc for doc in docs if doc.get("filename") == filename), None)
    if not match:
        raise HTTPException(404, "RAG document not found")

    doc_id = str(match.get("doc_id") or "")
    job_id = f"delete_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}"
    with upload_jobs_lock:
        upload_jobs[job_id] = {
            "job_id": job_id,
            "kind": "delete",
            "status": "queued",
            "stage": "queued",
            "message": "삭제 대기 중...",
            "progress": 5,
            "filename": filename,
            "doc_id": doc_id,
            "error": None,
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "updated_at": datetime.now().isoformat(timespec="seconds"),
        }

    worker = threading.Thread(target=_process_delete_job, args=(job_id, filename, doc_id), daemon=True)
    worker.start()
    return _job_snapshot(job_id)


@app.post("/api/rag/qa")
async def rag_qa(proxy_request: Request):
    payload = await proxy_request.json()
    if not isinstance(payload, dict):
        raise HTTPException(400, "Expected JSON object")
    query = str(payload.get("query") or "").strip()
    filename_filter = str(payload.get("filename_filter") or "").strip()
    if filename_filter and _is_document_title_query(query):
        return _document_title_response(query, filename_filter)
    cache_key = _qa_cache_key(payload)
    with rag_response_cache_lock:
        cached = qa_response_cache.get(cache_key)
    if cached is not None:
        return {**cached, "cache_hit": True}

    response = _rag_api_post("/qa", payload)
    with rag_response_cache_lock:
        qa_response_cache[cache_key] = response
    return {**response, "cache_hit": False}


@app.post("/api/rag/qa/stream")
async def rag_qa_stream(proxy_request: Request):
    payload = await proxy_request.json()
    if not isinstance(payload, dict):
        raise HTTPException(400, "Expected JSON object")
    query = str(payload.get("query") or "").strip()
    filename_filter = str(payload.get("filename_filter") or "").strip()
    if filename_filter and _is_document_title_query(query):
        response = _document_title_response(query, filename_filter)
        with rag_response_cache_lock:
            qa_response_cache[_qa_cache_key(payload)] = response
        return StreamingResponse(
            _stream_cached_qa_response(response),
            media_type="text/event-stream",
            headers={"Cache-Control": "no-cache", "Connection": "keep-alive", "X-Accel-Buffering": "no"},
        )

    cache_key = _qa_cache_key(payload)
    with rag_response_cache_lock:
        cached = qa_response_cache.get(cache_key)
    if cached is not None:
        return StreamingResponse(
            _stream_cached_qa_response(cached),
            media_type="text/event-stream",
            headers={"Cache-Control": "no-cache", "Connection": "keep-alive", "X-Accel-Buffering": "no"},
        )

    return StreamingResponse(
        _rag_api_stream("/qa/stream", payload, cache_key=cache_key),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "Connection": "keep-alive", "X-Accel-Buffering": "no"},
    )


@app.post("/api/rag/summary")
async def rag_summary(proxy_request: Request):
    payload = await proxy_request.json()
    if not isinstance(payload, dict):
        raise HTTPException(400, "Expected JSON object")
    cache_key = _summary_cache_key(payload)
    with rag_response_cache_lock:
        cached = summary_response_cache.get(cache_key)
    if cached is not None:
        return {**cached, "cache_hit": True}

    response = _rag_api_post("/summary", payload)
    with rag_response_cache_lock:
        summary_response_cache[cache_key] = response
    return {**response, "cache_hit": False}


@app.post("/api/documents/{doc_id}/parse")
async def parse_single(doc_id: str):
    for fname in os.listdir(DOC_DIR):
        ext = os.path.splitext(fname)[1].lower()
        if ext not in SUPPORTED_EXTENSIONS:
            continue
        fpath = os.path.join(DOC_DIR, fname)
        if _doc_id(fpath) == doc_id:
            try:
                result = parse_document(fpath)
            except Exception as exc:
                logger.exception("parse_document failed for %s", doc_id)
                raise HTTPException(500, f"Parse failed: {exc}") from exc
            if result.get("status") == "error":
                return result
            _save_result(doc_id, result)
            try:
                write_root_llm_exports(result, PROJECT_ROOT)
            except Exception as exc:
                logger.warning("write_root_llm_exports failed: %s", exc)
            parsed_cache[doc_id] = result
            _maybe_create_sidecar_pdf(result)
            return result
    raise HTTPException(404, "Document not found")


@app.get("/api/documents/{doc_id}")
async def get_parsed(doc_id: str):
    cached = parsed_cache.get(doc_id)
    if cached is None:
        cached = _load_result(doc_id)
        if cached:
            parsed_cache[doc_id] = cached
    if cached is None:
        raise HTTPException(404, "Document not parsed yet")
    return cached


@app.get("/api/documents/{doc_id}/viewer")
async def get_document_viewer(doc_id: str):
    cached = parsed_cache.get(doc_id)
    if cached is None:
        cached = _load_result(doc_id)
        if cached:
            parsed_cache[doc_id] = cached
    if cached is None:
        raise HTTPException(404, "Document not parsed yet")
    return _viewer_payload_with_sidecar(cached)


@app.post("/api/parse-all")
async def parse_all():
    cleanup_generated_artifacts(
        PROJECT_ROOT,
        parsed_results_dir=Path(RESULT_DIR),
    )
    parsed_cache.clear()
    results: list[dict[str, Any]] = []
    for fname in sorted(os.listdir(DOC_DIR)):
        ext = os.path.splitext(fname)[1].lower()
        if ext not in SUPPORTED_EXTENSIONS:
            continue
        fpath = os.path.join(DOC_DIR, fname)
        if not os.path.isfile(fpath):
            continue
        did = _doc_id(fpath)
        try:
            result = parse_document(fpath)
        except Exception as exc:
            result = {
                "id": did,
                "filename": fname,
                "status": "error",
                "error": str(exc),
                "pages": [],
                "metadata": {},
            }
        parsed_cache[did] = result
        _save_result(did, result)
        if result.get("status") != "error":
            _maybe_create_sidecar_pdf(result)
        results.append(result)
    return results


@app.get("/prototype-summary", response_class=HTMLResponse)
async def prototype_summary_page():
    return FileResponse(
        os.path.join(os.path.dirname(__file__), "static", "summary.html"),
        media_type="text/html",
    )


@app.get("/api/prototype-stats")
async def prototype_stats():
    harness_path = os.path.join(os.path.dirname(__file__), "harness_history.json")
    if not os.path.exists(harness_path):
        raise HTTPException(404, "Harness history not found")
    try:
        with open(harness_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as exc:
        raise HTTPException(500, f"Error reading harness history: {exc}")


@app.get("/api/quality-report")
async def quality_report():
    report_path = os.path.join(os.path.dirname(__file__), "quality_report.md")
    if not os.path.exists(report_path):
        raise HTTPException(404, "Quality report not found")
    try:
        with open(report_path, "r", encoding="utf-8") as f:
            return {"content": f.read()}
    except Exception as exc:
        raise HTTPException(500, f"Error reading quality report: {exc}")


# ── preview image serving ────────────────────────────────────────────

@app.get("/api/documents/{doc_id}/pages/{page_num}/preview")
async def get_page_preview(doc_id: str, page_num: int):
    """Serve a rendered page preview image (PNG)."""
    preview_path = _ensure_pdf_preview(doc_id, page_num) or _preview_path(doc_id, page_num)
    if not preview_path.is_file():
        raise HTTPException(404, "Preview image not found. Parse the document first.")
    return FileResponse(preview_path, media_type="image/png")


@app.get("/api/documents/{doc_id}/preview-strip")
async def get_spreadsheet_preview_strip(doc_id: str):
    """Serve a stitched spreadsheet preview image (PNG)."""
    preview_path = _ensure_spreadsheet_stitched_preview(doc_id)
    if preview_path is None or not preview_path.is_file():
        raise HTTPException(404, "Stitched preview image not found. Open the document again after parsing.")
    return FileResponse(preview_path, media_type="image/png")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=int(os.getenv("APP_PORT", "8001")))
