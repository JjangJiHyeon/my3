from __future__ import annotations

from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment


ROOT = Path(r"C:\Users\jihyeon\Desktop\my3")
TEMPLATE_DIR = ROOT / "_templates"
OUT_FEATURE = ROOT / "기능정의서_my3.xlsx"
OUT_DESIGN = ROOT / "개발설계서_my3.xlsx"


def set_value(ws, coord: str, value: str) -> None:
    ws[coord] = value
    alignment = copy(ws[coord].alignment)
    alignment.wrap_text = True
    ws[coord].alignment = alignment


def set_row_height(ws, row: int, height: float) -> None:
    ws.row_dimensions[row].height = height


def copy_row_style(ws, src_row: int, dst_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.protection:
            dst.protection = copy(src.protection)
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def merge_row(ws, row: int, start_col: int, end_col: int) -> None:
    ws.merge_cells(
        start_row=row,
        start_column=start_col,
        end_row=row,
        end_column=end_col,
    )


def ensure_row_merge(ws, row: int, start_col: int, end_col: int) -> None:
    target = f"{openpyxl.utils.get_column_letter(start_col)}{row}:{openpyxl.utils.get_column_letter(end_col)}{row}"
    if target not in {str(rng) for rng in ws.merged_cells.ranges}:
        ws.merge_cells(target)


def append_component_rows(ws) -> None:
    for new_row in (19, 20, 21):
        copy_row_style(ws, 18, new_row)
        ensure_row_merge(ws, new_row, 3, 6)


def append_workflow_rows(ws) -> None:
    for new_row in (13, 14):
        copy_row_style(ws, 12, new_row)
        ensure_row_merge(ws, new_row, 3, 6)


def prepare_tech_sheet(ws) -> None:
    if "E6:F11" in {str(rng) for rng in ws.merged_cells.ranges}:
        ws.unmerge_cells("E6:F11")
    for row in range(6, 12):
        ensure_row_merge(ws, row, 5, 6)
        copy_row_style(ws, 12, row)
    if "B20:F20" in {str(rng) for rng in ws.merged_cells.ranges}:
        ws.unmerge_cells("B20:F20")
    copy_row_style(ws, 21, 20)


def left_top_wrap(ws, coord: str) -> None:
    cell = ws[coord]
    cell.alignment = Alignment(
        horizontal=cell.alignment.horizontal or "left",
        vertical="top",
        wrap_text=True,
    )


def fill_feature_workbook(path: Path) -> None:
    wb = openpyxl.load_workbook(path)

    ws = wb["프로젝트 개요"]
    set_value(ws, "B2", "my3 기능 정의서")
    set_value(ws, "C5", "my3")
    set_value(ws, "C6", "2026-04-08")
    set_value(
        ws,
        "B9",
        "지원 확장자(.pdf, .doc, .docx, .hwp, .xls, .xlsx) 문서를 파싱해 parsed_results, review JSON, preview 이미지, "
        "latest_sidecars LLM sidecar를 만든다.\n"
        "이후 chunk JSON 생성 -> OpenAI 임베딩 기반 vector_indexes 생성 -> Chroma 적재를 수행하고, "
        "정적 UI와 FastAPI API에서 문서 선택형 QA/문서 전체 요약을 제공한다.\n"
        "메인 앱(app.py)은 업로드/삭제/UI/프록시를 담당하고, 실제 QA·summary 응답은 별도 RAG API(scripts/run_rag_api.py)가 처리한다.",
    )
    set_value(ws, "B13", "Step 1")
    set_value(ws, "C13", "문서 업로드/관리")
    set_value(
        ws,
        "D13",
        "정적 UI와 /api/rag/upload, /api/rag/delete가 documents/에 파일을 저장·삭제하고 job_id, stage, progress 기반 작업 상태를 제공한다.",
    )
    set_value(ws, "B14", "Step 2")
    set_value(ws, "C14", "형식별 파싱")
    set_value(
        ws,
        "D14",
        "parse_document가 PDF/DOC/DOCX/HWP/XLS/XLSX parser를 호출해 parsed_results/{doc_id}.json, review JSON, page preview, "
        "latest_sidecars llm_ready/md를 생성한다.",
    )
    set_value(ws, "B15", "Step 3")
    set_value(ws, "C15", "청킹·벡터 준비")
    set_value(
        ws,
        "D15",
        "scripts.build_chunks가 전략별 chunk JSON을 만들고, scripts.run_retrieval_prep가 metadata JSONL, vectors.npz, id_map, run_manifest를 생성한다.",
    )
    set_value(ws, "B16", "Step 4")
    set_value(ws, "C16", "Chroma 적재·QA/요약")
    set_value(
        ws,
        "D16",
        "scripts.run_chroma_ingest가 Chroma persistent collection(document_chunks)을 갱신하고, "
        "/api/rag/qa·/api/rag/summary가 검색 결과로 gpt-5.2 응답과 source(문서/페이지)를 반환한다.",
    )
    set_value(
        ws,
        "B19",
        "  • RAG 응답 경로는 app.py(기본 8001)와 별도 RAG API(RAG_API_BASE_URL, 기본 http://127.0.0.1:8000) 두 프로세스가 연동되어야 한다.",
    )
    set_value(
        ws,
        "B20",
        "  • OPENAI_API_KEY는 환경 변수 또는 .env에 필요하며, 코드가 허용하는 모델/차원은 gpt-5.2 / text-embedding-3-large / 3072로 고정되어 있다.",
    )
    set_value(
        ws,
        "B21",
        "  • PDF가 table_heavy로 라우팅되더라도 현재 전용 pipeline은 없어서 text_report_pipeline으로 fallback한다.",
    )
    set_value(
        ws,
        "B22",
        "  • .doc 파싱은 Tika -> win32com -> OLE binary 순으로 시도하며, win32com 경로는 pywin32와 Microsoft Word가 있는 환경에서만 동작한다.",
    )

    for coord in ("B9", "D13", "D14", "D15", "D16", "B19", "B20", "B21", "B22"):
        left_top_wrap(ws, coord)
    set_row_height(ws, 9, 94)
    for row in range(13, 17):
        set_row_height(ws, row, 48)
    for row in range(19, 23):
        set_row_height(ws, row, 42)

    ws = wb["요건 정의 상세"]
    set_value(ws, "B2", "요건 정의 상세  |  my3")
    set_value(ws, "B5", "문서 관리\n및 UI")
    set_value(ws, "B9", "파싱 및\n기초 산출물")
    set_value(ws, "B14", "청킹·벡터\n준비/적재")
    set_value(ws, "B19", "RAG 응답\n및 추적")

    requirement_rows = {
        5: (
            "REQ-M-01",
            "RAG 문서 목록 조회",
            "최신 Chroma 색인의 id_manifest.json을 파일명 기준으로 집계해 filename, doc_id, chunk_count, page_count, strategies를 반환한다. "
            "UI의 문서 목록과 전략 선택값이 이를 사용한다.",
            "app.py::_load_rag_documents\napp.py::/api/rag/documents\nstatic/index.html::loadDocuments",
            "chroma_indexes/latest.json 또는 id_manifest.json이 없으면 빈 목록 반환",
        ),
        6: (
            "REQ-M-02",
            "문서 업로드 및 비동기 인덱싱",
            "multipart/form-data file을 documents/에 저장한 뒤 백그라운드 스레드에서 parse -> chunk -> embedding -> chroma 순서로 인덱싱한다. "
            "응답은 job_id, kind, status, stage, progress, filename 등을 가진 작업 스냅샷이다.",
            "app.py::/api/rag/upload\napp.py::_process_upload_job\nstatic/index.html::uploadSelectedFile",
            "허용 확장자: .pdf/.doc/.docx/.hwp/.xls/.xlsx\nindex_build_lock으로 색인 작업 직렬화",
        ),
        7: (
            "REQ-M-03",
            "업로드/삭제 작업 상태 조회",
            "업로드와 삭제 작업 상태를 조회한다. UI가 약 1.8초 간격으로 poll하여 progress bar와 메시지를 갱신한다.",
            "app.py::/api/rag/upload-jobs/{job_id}\nstatic/index.html::pollUploadJob",
            "존재하지 않는 job_id는 404",
        ),
        8: (
            "REQ-M-04",
            "문서 삭제 및 롤백",
            "선택 문서와 parsed_results/chunks 산출물을 delete_trash로 안전 이동한 뒤, 최신 vector index에서 해당 doc_id만 제외한 filtered run을 만들고 "
            "Chroma를 재생성한다. 실패 시 latest.json과 파일 이동을 복구한다.",
            "app.py::/api/rag/delete\napp.py::_process_delete_job\napp.py::_create_filtered_vector_run_excluding_doc",
            "문서명 기준 삭제 요청\nRAG 문서 목록에 없는 filename은 404",
        ),
        9: (
            "REQ-P-01",
            "멀티포맷 파싱 라우팅",
            "parse_document가 확장자별 parser를 dispatch하여 공통 결과 구조(id, filename, filepath, file_type, file_size, pages, metadata, status, parser_version)를 만든다.",
            "parsers/__init__.py::parse_document\nSUPPORTED_EXTENSIONS",
            "지원 외 형식은 status=error 반환",
        ),
        10: (
            "REQ-P-02",
            "PDF 파싱 및 페이지 미리보기",
            "PDF는 pre_route_document로 초기 문서 유형을 추정한 뒤 slide_ir/text_report/dashboard_brief pipeline 중 하나를 선택한다. "
            "페이지별 block, tables, rag_text, parser_debug, preview_image를 만들고 route_document 결과를 metadata에 기록한다.",
            "parsers/pdf_parser.py\nparsers/document_router.py\nparsers/pdf/*",
            "table_heavy는 현재 text_report_pipeline으로 fallback",
        ),
        11: (
            "REQ-P-03",
            "DOC/DOCX/HWP 파싱",
            "DOCX는 python-docx, DOC는 Tika -> win32com -> OLE binary 순으로 시도하고, HWP는 custom OLE parser로 BodyText/Section을 읽는다. "
            "두 parser 모두 route_document 후 페이지별 rag_text를 생성한다.",
            "parsers/doc_parser.py\nparsers/hwp_parser.py",
            ".doc win32com 경로는 pywin32와 Microsoft Word 필요",
        ),
        12: (
            "REQ-P-04",
            "XLS/XLSX 파싱 및 시트 페이지화",
            "XLSX는 openpyxl, XLS는 xlrd로 읽는다. 각 시트를 page로 변환해 text, tables, row_count, col_count를 반환한다.",
            "parsers/xlsx_parser.py",
            "preview 이미지/OCR 없이 시트 데이터만 구조화",
        ),
        13: (
            "REQ-P-05",
            "파싱 결과 저장 및 sidecar 생성",
            "파싱 성공 결과를 parsed_results/{doc_id}.json과 parsed_results/review/{doc_id}.json에 저장하고, "
            "exports/latest_sidecars에 llm_ready JSON과 md report를 함께 생성한다.",
            "app_support/review_export.py\napp_support/export_to_gpt.py\napp.py::_save_result",
            "review/sidecar는 보조 산출물이며 판단 기준은 실제 파서/스크립트 흐름",
        ),
        14: (
            "REQ-V-01",
            "본문 우선 청크 생성",
            "text_first_with_visual_support 전략은 text block을 panel/flow 단위로 묶고, 관련 table/chart/image 요약을 support text로 붙여 retrieval_text를 만든다.",
            "chunking/strategies.py::build_text_first_chunks",
            "chunks/{doc_id}.text_first_with_visual_support.json",
        ),
        15: (
            "REQ-V-02",
            "LLM-ready baseline 청크 생성",
            "llm_ready_native 전략은 exports/latest_sidecars의 page rag_text와 block summary를 기반으로 page/block 단위 청크를 만든다. "
            "llm_ready sidecar가 없으면 issue를 기록한다.",
            "chunking/strategies.py::build_llm_ready_native_chunks\nchunking/io.py::load_llm_ready",
            "chunks/{doc_id}.llm_ready_native.json",
        ),
        16: (
            "REQ-V-03",
            "청크 요약 및 retrieval metadata 정규화",
            "문서별 chunk_summary.json을 만들고, 각 chunk를 vector_record_id, page_num, chunk_type, retrieval_text, evidence_preview 등 retrieval metadata 계약으로 정규화한다.",
            "chunking/builders.py\nretrieval/metadata_builder.py\nretrieval/metadata_schema.py",
            "vector_indexes/*/chunk_metadata_records.jsonl",
        ),
        17: (
            "REQ-V-04",
            "OpenAI 임베딩 및 vector 파일 생성",
            "비어 있지 않은 retrieval_text만 OpenAI 임베딩으로 변환해 vectors.npz와 id_map.json을 만든다. "
            "모델은 text-embedding-3-large, 차원은 3072로 고정이다.",
            "retrieval/vector_prep.py\nretrieval/embedding_runner.py",
            "OPENAI_API_KEY 필요\nvector_indexes/latest.json 갱신",
        ),
        18: (
            "REQ-V-05",
            "Chroma 적재 및 latest 갱신",
            "준비된 벡터/메타데이터를 Chroma persistent collection(document_chunks)에 upsert하고 run_manifest, ingest_summary, id_manifest를 기록한다.",
            "chroma_store/ingest.py\nchroma_store/collection_builder.py\nscripts/run_chroma_ingest.py",
            "chroma_indexes/latest.json 갱신\nbatch_size 기본 256",
        ),
        19: (
            "REQ-R-01",
            "문서 기반 QA",
            "query, strategy_name, top_k, filename_filter를 받아 선택 문서 기준 유사 청크를 검색하고 ChatOpenAI(gpt-5.2)로 한국어 답변을 생성한다.",
            "app.py::/api/rag/qa\nscripts/run_rag_api.py::/qa\nrag_api/retriever.py\nrag_api/qa_chain.py",
            "filename_filter가 있고 제목 질의면 app.py가 파일명 응답을 직접 반환\nQA top_k 1~30",
        ),
        20: (
            "REQ-R-02",
            "문서 전체 요약",
            "filename, strategy_name, top_k를 받아 문서 전체 요약용 coverage chunk를 페이지 분포 기준으로 선택하고 고정 섹션 형식의 요약을 생성한다.",
            "app.py::/api/rag/summary\nscripts/run_rag_api.py::/summary\nrag_api/retriever.py::retrieve_representative\nrag_api/summary_chain.py",
            "Summary top_k 1~50\nsource는 문서/페이지 단위로 dedupe",
        ),
        21: (
            "REQ-R-03",
            "결과·출처·전략 표시",
            "정적 UI가 문서 선택, drag/drop, strategy 선택, top_k 선택, 업로드/삭제 progress, QA/summary 결과, source chip(문서·page)를 표시한다.",
            "static/index.html\napp.py::/\napp.py::/api/rag/documents",
            "선택 문서가 없으면 QA/summary 실행 불가",
        ),
        22: (
            "REQ-R-04",
            "보조 조회 및 실행 추적",
            "preview 이미지 조회, prototype stats/quality report 조회, rag_api_runs의 run_manifest·latency_events 기록으로 보조 검증과 추적을 제공한다.",
            "app.py::/api/documents/{doc_id}/pages/{page_num}/preview\napp.py::/api/prototype-stats\napp.py::/api/quality-report\nrag_api/observability.py",
            "quality_report.md 또는 harness_history.json이 없으면 404",
        ),
    }
    for row, values in requirement_rows.items():
        set_value(ws, f"C{row}", values[0])
        set_value(ws, f"D{row}", values[1])
        set_value(ws, f"E{row}", values[2])
        set_value(ws, f"F{row}", values[3])
        set_value(ws, f"G{row}", values[4])
        for coord in (f"D{row}", f"E{row}", f"F{row}", f"G{row}"):
            left_top_wrap(ws, coord)
        set_row_height(ws, row, 84)

    ws = wb["Input 및 Output 정의"]
    set_value(ws, "B2", "Input / Output 정의 | my3")
    set_value(ws, "B5", "Input")
    set_value(
        ws,
        "C5",
        "1. 업로드 입력\n"
        "- multipart/form-data file (.pdf, .doc, .docx, .hwp, .xls, .xlsx)\n\n"
        "2. 삭제 요청\n"
        '- JSON { "filename": "<문서명>" }\n\n'
        "3. QA 요청\n"
        '- JSON { "query": str, "strategy_name": str, "top_k": int, "filename_filter": str | null }\n\n'
        "4. Summary 요청\n"
        '- JSON { "filename": str, "strategy_name": str, "top_k": int }\n\n'
        "5. 관리 API 입력\n"
        "- /api/documents/{doc_id}/parse\n"
        "- /api/parse-all\n\n"
        "6. 배치/CLI 입력\n"
        "- scripts.build_chunks: --from-cache | --doc-id | --file, --out-dir\n"
        "- scripts.run_retrieval_prep: --chunks-dir, --output-dir, --embedding-batch-size\n"
        "- scripts.run_chroma_ingest: --vector-run, --persist-dir, --collection-name, --batch-size\n\n"
        "7. 환경 설정\n"
        "- .env / 환경변수: OPENAI_API_KEY, OPENAI_CHAT_MODEL, OPENAI_EMBEDDING_MODEL, OPENAI_EMBEDDING_DIMENSIONS,\n"
        "  CHROMA_COLLECTION_NAME, CHROMA_PERSIST_DIR\n"
        "- 선택적 LangSmith 변수: LANGSMITH_TRACING, LANGSMITH_API_KEY, LANGSMITH_PROJECT, LANGSMITH_ENDPOINT",
    )
    set_value(ws, "B6", "Output")
    set_value(
        ws,
        "C6",
        "1. 파싱 결과\n"
        "- parsed_results/{doc_id}.json\n"
        "- parsed_results/review/{doc_id}.json\n"
        "- parsed_results/previews/{doc_id}/page_{n}.png\n\n"
        "2. LLM sidecar / chunk 결과\n"
        "- exports/latest_sidecars/{stem}_llm_ready.json\n"
        "- exports/latest_sidecars/{stem}_llm_report.md\n"
        "- chunks/{doc_id}.text_first_with_visual_support.json\n"
        "- chunks/{doc_id}.llm_ready_native.json\n"
        "- chunks/{doc_id}.chunk_summary.json\n\n"
        "3. 벡터 준비 결과\n"
        "- vector_indexes/run_*/run_manifest.json\n"
        "- vector_indexes/run_*/chunk_metadata_records.jsonl\n"
        "- vector_indexes/run_*/embedding_stats.json\n"
        "- vector_indexes/run_*/vectors.npz\n"
        "- vector_indexes/run_*/id_map.json\n"
        "- vector_indexes/latest.json\n\n"
        "4. Chroma 결과\n"
        "- chroma_indexes/run_*/run_manifest.json\n"
        "- chroma_indexes/run_*/ingest_summary.json\n"
        "- chroma_indexes/run_*/id_manifest.json\n"
        "- chroma_indexes/run_*/chroma/\n"
        "- chroma_indexes/latest.json\n\n"
        "5. API 응답\n"
        "- 업로드/삭제 job snapshot: {job_id, kind, status, stage, progress, message, filename, doc_id, error, created_at, updated_at}\n"
        "- RAG response: {mode, title, answer, sources[{document, page}]}\n"
        "- 문서 목록: {filename, doc_id, chunk_count, page_count, strategies}\n\n"
        "6. 실행 추적\n"
        "- rag_api_runs/run_*/run_manifest.json\n"
        "- rag_api_runs/run_*/latency_events.jsonl\n"
        "- rag_api_runs/run_*/qa_samples.jsonl\n"
        "- rag_api_runs/run_*/summary_samples.jsonl",
    )
    left_top_wrap(ws, "C5")
    left_top_wrap(ws, "C6")
    set_row_height(ws, 5, 300)
    set_row_height(ws, 6, 360)

    wb.save(OUT_FEATURE)


def fill_design_workbook(path: Path) -> None:
    wb = openpyxl.load_workbook(path)

    ws = wb["프로그램 개요"]
    set_value(ws, "B2", "my3 개발 설계서")
    set_value(ws, "C5", "my3")
    set_value(ws, "C6", "2026-04-08")
    set_value(
        ws,
        "B9",
        "documents/의 멀티포맷 문서를 파싱하고, 결과를 parsed_results -> chunks -> vector_indexes -> chroma_indexes로 단계적으로 변환해 "
        "정적 UI와 별도 RAG API에서 QA/summary에 사용하는 파일 기반 파이프라인이다.\n"
        "메인 앱(app.py)은 업로드/삭제/UI 서빙과 RAG API 프록시를 담당하고, 색인 구축은 scripts/*.py, "
        "검색·응답은 rag_api/*.py로 분리되어 동작한다.",
    )
    set_value(
        ws,
        "C13",
        "진입점은 app.py이며 /, /api/documents, /api/rag/*, preview/quality/prototype 조회를 제공한다.\n"
        "parsers/는 parse_document dispatcher와 pdf/doc/hwp/xlsx parser, document_router, OCR/table 유틸을 포함한다.\n"
        "app_support/는 parsed_results/review JSON 저장과 latest llm_ready sidecar 생성을 담당한다.\n"
        "chunking/ -> retrieval/ -> chroma_store/ 순서로 chunk JSON, metadata/vector, Chroma persistent index를 생성한다.\n"
        "scripts/run_rag_api.py + rag_api/는 Chroma 검색, summary coverage selection, ChatOpenAI 응답을 수행한다.\n"
        "static/index.html은 문서 목록/업로드/삭제/QA/summary UI를 제공한다.",
    )
    set_value(
        ws,
        "C14",
        "업로드 흐름: /api/rag/upload -> _process_upload_job -> parse_document -> save_all_parse_outputs -> "
        "write_root_llm_exports -> scripts.build_chunks -> scripts.run_retrieval_prep -> scripts.run_chroma_ingest -> latest.json 갱신\n"
        "조회/질의 흐름: UI 문서 선택 -> /api/rag/summary 또는 /api/rag/qa -> app.py proxy -> run_rag_api /summary 또는 /qa -> "
        "ChromaRetriever(strategy/filename filter) -> summary_chain 또는 qa_chain -> answer + sources 반환\n"
        "삭제 흐름: /api/rag/delete -> delete_trash staging -> filtered vector run 생성 -> run_chroma_ingest --vector-run <filtered> -> "
        "성공 시 trash 정리 / 실패 시 latest 및 파일 복구",
    )
    for coord in ("B9", "C13", "C14"):
        left_top_wrap(ws, coord)
    set_row_height(ws, 9, 86)
    set_row_height(ws, 13, 168)
    set_row_height(ws, 14, 150)

    ws = wb["기술 스택 및 미결사항"]
    prepare_tech_sheet(ws)
    set_value(ws, "B2", "기술 스택 요약  &  미결 사항 추적 | my3")
    set_value(ws, "B6", "언어·런타임")
    set_value(ws, "C6", "Python 3.11")
    set_value(ws, "D6", "app.py, scripts, parser/chunking/retrieval/chroma/rag_api 전체 구현 런타임")
    set_value(ws, "E6", "현재 실행 환경과 pyc 경로 기준")

    set_value(ws, "B7", "웹/API\n및 UI")
    set_value(ws, "C7", "FastAPI, Uvicorn, python-multipart")
    set_value(ws, "D7", "app.py(메인 API)와 scripts/run_rag_api.py(RAG API), 파일 업로드 처리")
    set_value(ws, "E7", "메인 앱 기본 포트 8001, RAG API 기본 포트 8000")
    set_value(ws, "C8", "HTML/CSS/JavaScript (static/index.html, static/summary.html)")
    set_value(ws, "D8", "문서 목록/업로드/삭제/QA/summary UI와 prototype 화면")
    set_value(ws, "E8", "프론트 빌드 도구 없이 정적 파일 서빙")

    set_value(ws, "B9", "PDF 파싱")
    set_value(ws, "C9", "PyMuPDF(fitz), pdfplumber, camelot-py[cv], pymupdf4llm")
    set_value(ws, "D9", "페이지 텍스트·block·table·preview 추출과 pipeline별 PDF 처리")
    set_value(ws, "E9", "text_report / slide_ir / dashboard_brief pipeline 사용")

    set_value(ws, "B10", "문서 파싱")
    set_value(ws, "C10", "python-docx, tika, olefile")
    set_value(ws, "D10", "DOCX 파싱과 DOC/HWP 계열 텍스트 추출 전략")
    set_value(ws, "E10", ".doc는 Tika/COM/OLE 순서, HWP는 custom OLE parser")
    set_value(ws, "C11", "openpyxl, xlrd")
    set_value(ws, "D11", "XLSX/XLS 시트별 text/table 변환")
    set_value(ws, "E11", "시트가 page 단위 결과로 저장됨")

    set_value(ws, "B12", "OCR")
    set_value(ws, "C12", "easyocr, pytesseract, opencv-python, Pillow, numpy")
    set_value(ws, "D12", "선택적 OCR과 이미지 전처리")
    set_value(ws, "E12", "EasyOCR 우선, pytesseract fallback")

    set_value(ws, "B13", "청킹")
    set_value(ws, "C13", "chunking/* (text_first_with_visual_support, llm_ready_native)")
    set_value(ws, "D13", "parsed result/llm_ready를 chunk JSON으로 변환")
    set_value(ws, "E13", "support text, page/block metadata 포함")

    set_value(ws, "B14", "임베딩·벡터 준비")
    set_value(ws, "C14", "langchain-openai, openai fallback, numpy")
    set_value(ws, "D14", "retrieval_text 임베딩과 vector_indexes 산출물 생성")
    set_value(ws, "E14", "모델 고정: text-embedding-3-large, 3072 dims")

    set_value(ws, "B15", "벡터 저장소\n및 RAG")
    set_value(ws, "C15", "chromadb, langchain-chroma")
    set_value(ws, "D15", "Chroma persistent collection(document_chunks) 구축 및 조회")
    set_value(ws, "E15", "id_manifest.json, latest.json 관리")
    set_value(ws, "C16", "langchain-core, ChatOpenAI(gpt-5.2), pydantic, optional LangSmith")
    set_value(ws, "D16", "QA/summary chain, 요청/응답 schema, 선택적 tracing")
    set_value(ws, "E16", "LangSmith는 env와 패키지가 있을 때만 활성화")

    set_value(ws, "B20", "PDF table_heavy 전용 파이프라인")
    set_value(
        ws,
        "C20",
        "document_router가 table_heavy로 분류해도 parse_pdf는 text_report_pipeline으로 fallback한다. metadata에 pipeline_fallback_reason이 남는다.",
    )
    set_value(ws, "D20", "")
    set_value(ws, "E20", "")
    set_value(ws, "F20", "미결")
    set_value(ws, "B21", ".doc win32com 의존성")
    set_value(
        ws,
        "C21",
        "win32com 전략은 pywin32와 Microsoft Word 설치가 있어야 하지만 optional이다. 실패하면 Tika/OLE 전략으로 넘어간다.",
    )
    set_value(ws, "D21", "")
    set_value(ws, "E21", "")
    set_value(ws, "F21", "조건부")
    set_value(ws, "B22", "별도 RAG API 기동 의존성")
    set_value(
        ws,
        "C22",
        "app.py의 QA/summary는 RAG_API_BASE_URL(기본 http://127.0.0.1:8000)로 프록시한다. 미기동 시 502를 반환한다.",
    )
    set_value(ws, "D22", "")
    set_value(ws, "E22", "")
    set_value(ws, "F22", "운영 제약")

    for coord in (
        "C6",
        "D6",
        "E6",
        "C7",
        "D7",
        "E7",
        "C8",
        "D8",
        "E8",
        "C9",
        "D9",
        "E9",
        "C10",
        "D10",
        "E10",
        "C11",
        "D11",
        "E11",
        "C12",
        "D12",
        "E12",
        "C13",
        "D13",
        "E13",
        "C14",
        "D14",
        "E14",
        "C15",
        "D15",
        "E15",
        "C16",
        "D16",
        "E16",
        "C20",
        "C21",
        "C22",
    ):
        left_top_wrap(ws, coord)
    for row in range(6, 17):
        set_row_height(ws, row, 44)
    for row in range(20, 23):
        set_row_height(ws, row, 48)

    ws = wb["프로그램 구성안"]
    append_component_rows(ws)
    set_value(ws, "B2", "프로그램 구성안 | my3")
    set_value(
        ws,
        "B5",
        "멀티포맷 문서를 파싱하고 parsed_results -> chunks -> vector_indexes -> chroma_indexes로 단계별 산출물을 생성한 뒤, "
        "정적 UI와 별도 RAG API에서 QA/summary를 제공하는 파일 기반 구조이다.",
    )
    components = {
        13: (
            "app.py",
            "메인 FastAPI 진입점. /, /api/documents, /api/rag/*, preview/quality/prototype 조회를 제공하고, "
            "업로드/삭제 작업 스레드와 parse -> chunk -> embedding -> chroma 실행, RAG API 프록시를 담당한다.",
        ),
        14: (
            "app_support/",
            "parse 결과 저장·정리 보조 모듈. review JSON 저장(review_export.py), latest llm_ready/md sidecar 생성(export_to_gpt.py), "
            "generated artifact cleanup를 담당한다.",
        ),
        15: (
            "parsers/",
            "멀티포맷 파서 계층. parse_document dispatcher와 pdf/doc/hwp/xlsx parser, document_router, OCR/table 유틸, PDF pipeline이 포함된다. "
            "입력은 documents 파일, 출력은 pages/blocks/metadata/rag_text 구조이다.",
        ),
        16: (
            "chunking/",
            "parsed 결과와 llm_ready sidecar를 chunk JSON으로 바꾸는 계층. main(text_first_with_visual_support)과 "
            "baseline(llm_ready_native) 전략, visual support/summary 유틸을 포함한다.",
        ),
        17: (
            "retrieval/",
            "chunk JSON을 retrieval metadata와 벡터 파일로 정규화하는 계층. chunk_metadata_records.jsonl, vectors.npz, id_map.json, "
            "run_manifest.json을 생성한다.",
        ),
        18: (
            "chroma_store/",
            "vector_indexes 산출물을 Chroma persistent collection으로 적재하는 계층. collection builder, metadata scalar coercion, "
            "ingest summary/id manifest/latest update를 담당한다.",
        ),
        19: (
            "rag_api/",
            "검색·응답 계층. config/observability/retriever/qa_chain/summary_chain/source_formatter가 Chroma 검색, coverage selection, "
            "ChatOpenAI 호출, source formatting을 수행한다.",
        ),
        20: (
            "scripts/",
            "배치 및 실행 스크립트 계층. parse_all/reparse_all, build_chunks, run_retrieval_prep, run_chroma_ingest, run_rag_api가 "
            "파이프라인 단계 실행을 캡슐화한다.",
        ),
        21: (
            "static/",
            "정적 UI 계층. index.html이 문서 목록/업로드/삭제/QA/summary 상호작용을 제공하고, summary.html이 "
            "prototype stats/quality report 조회 화면을 제공한다.",
        ),
    }
    for row, (name, desc) in components.items():
        set_value(ws, f"B{row}", name)
        set_value(ws, f"C{row}", desc)
        left_top_wrap(ws, f"C{row}")
        set_row_height(ws, row, 54)
    left_top_wrap(ws, "B5")
    set_row_height(ws, 5, 70)

    ws = wb["프로그램 워크플로우"]
    append_workflow_rows(ws)
    set_value(ws, "B2", "프로그램 워크플로우 | my3")
    set_value(
        ws,
        "C6",
        "1. 업로드/색인화 흐름\n"
        "- /api/rag/upload -> file save -> _process_upload_job\n"
        "- parse_document -> save_all_parse_outputs -> write_root_llm_exports\n"
        "- scripts.build_chunks -> scripts.run_retrieval_prep -> scripts.run_chroma_ingest\n"
        "- vector_indexes/latest.json, chroma_indexes/latest.json 갱신 -> UI progress 완료\n\n"
        "2. 삭제/조회 흐름\n"
        "- /api/rag/delete -> artifact stage(delete_trash) -> filtered vector run 생성 -> run_chroma_ingest --vector-run <filtered> -> 성공 시 trash 정리 / 실패 시 복구\n"
        "- /api/rag/documents -> id_manifest 집계 -> 문서 목록/전략/페이지 수 표시",
    )
    workflow_rows = {
        10: (
            "업로드 및 인덱싱",
            "multipart file 저장 -> parse_document -> parsed_results/review/sidecar 저장 -> build_chunks -> run_retrieval_prep -> run_chroma_ingest -> job status completed",
        ),
        11: (
            "삭제 및 색인 복구",
            "filename 매칭 -> 관련 파일 staging -> latest vector run에서 doc_id 제외 -> filtered vector run manifest 생성 -> Chroma 재적재 -> latest rollback/restore 처리",
        ),
        12: (
            "QA 요청",
            "static/index.html -> /api/rag/qa -> app.py proxy -> run_rag_api /qa -> ChromaRetriever.retrieve(strategy_name + filename_filter) -> qa_chain -> {mode, title, answer, sources}",
        ),
        13: (
            "Summary 요청",
            "문서 선택 -> /api/rag/summary -> run_rag_api /summary -> ChromaRetriever.retrieve_representative(filename) -> coverage selection/rerank -> summary_chain -> {mode, title, answer, sources}",
        ),
        14: (
            "보조 조회 및 추적",
            "preview endpoint, prototype stats, quality report, rag_api_runs/run_manifest.json, latency_events.jsonl, qa/summary sample jsonl로 결과 검증과 실행 추적을 지원",
        ),
    }
    for row, (name, desc) in workflow_rows.items():
        set_value(ws, f"B{row}", name)
        set_value(ws, f"C{row}", desc)
        left_top_wrap(ws, f"C{row}")
        set_row_height(ws, row, 52)
    left_top_wrap(ws, "C6")
    set_row_height(ws, 6, 190)

    wb.save(OUT_DESIGN)


def scan_for_leftover_examples(path: Path) -> list[str]:
    banned = [
        "문서 내용 요약 에이전트",
        "(예시)",
        "Streamlit",
        "Pinecone",
        "교육용 PDF 자료",
        "TXT 등",
        "후보",
    ]
    wb = openpyxl.load_workbook(path)
    hits: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str):
                    continue
                for token in banned:
                    if token in value:
                        hits.append(f"{path.name}:{ws.title}:{cell.coordinate}:{token}")
    return hits


def main() -> None:
    fill_feature_workbook(TEMPLATE_DIR / "feature_template.xlsx")
    fill_design_workbook(TEMPLATE_DIR / "design_template.xlsx")

    leftovers = scan_for_leftover_examples(OUT_FEATURE) + scan_for_leftover_examples(OUT_DESIGN)
    if leftovers:
        raise RuntimeError("Template example text remains:\n" + "\n".join(leftovers))


if __name__ == "__main__":
    main()
