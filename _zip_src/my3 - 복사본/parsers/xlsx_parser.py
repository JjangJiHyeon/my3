"""
Excel parser — openpyxl for .xlsx, xlrd for .xls.

Dates, numbers, and formula results are safely converted to strings.
Empty-only rows are excluded.
"""

from __future__ import annotations

import logging
import os
from datetime import datetime, date, time
from typing import Any

logger = logging.getLogger(__name__)


# ── public entry point ───────────────────────────────────────────────

def parse_excel(filepath: str) -> dict[str, Any]:
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".xls":
        return _parse_xls(filepath)
    return _parse_xlsx(filepath)


# ── .xlsx via openpyxl ───────────────────────────────────────────────

def _parse_xlsx(filepath: str) -> dict[str, Any]:
    import openpyxl

    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)

    pages: list[dict[str, Any]] = []
    for idx, name in enumerate(wb.sheetnames):
        ws = wb[name]
        rows_data = _collect_rows(ws.iter_rows())
        pages.append(_build_page(idx, name, rows_data))

    wb.close()

    metadata: dict[str, Any] = {
        "sheet_names": wb.sheetnames,
        "sheet_count": len(wb.sheetnames),
        "page_count": len(pages),
        "parser_used": "openpyxl (.xlsx)",
    }
    return {"pages": pages, "metadata": metadata, "status": "success"}


# ── .xls via xlrd ────────────────────────────────────────────────────

def _parse_xls(filepath: str) -> dict[str, Any]:
    import xlrd

    wb = xlrd.open_workbook(filepath, on_demand=True)

    pages: list[dict[str, Any]] = []
    for idx in range(wb.nsheets):
        ws = wb.sheet_by_index(idx)
        name = ws.name

        rows_data: list[list[str]] = []
        for r in range(ws.nrows):
            row_vals: list[str] = []
            for c in range(ws.ncols):
                cell = ws.cell(r, c)
                row_vals.append(_xlrd_cell_to_str(cell, wb))
            if any(v for v in row_vals):
                rows_data.append(row_vals)

        pages.append(_build_page(idx, name, rows_data))

    metadata: dict[str, Any] = {
        "sheet_names": [wb.sheet_by_index(i).name for i in range(wb.nsheets)],
        "sheet_count": wb.nsheets,
        "page_count": len(pages),
        "parser_used": "xlrd (.xls)",
    }
    return {"pages": pages, "metadata": metadata, "status": "success"}


# ── shared helpers ───────────────────────────────────────────────────

def _safe_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, (datetime, date, time)):
        return val.isoformat()
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return str(val)
    return str(val).strip()


def _collect_rows(row_iter: Any) -> list[list[str]]:
    rows: list[list[str]] = []
    for row in row_iter:
        vals = [_safe_str(cell.value) for cell in row]
        if any(v for v in vals):
            rows.append(vals)
    return rows


def _build_page(
    idx: int, sheet_name: str, rows_data: list[list[str]]
) -> dict[str, Any]:
    text_lines = ["\t".join(r) for r in rows_data if any(r)]
    return {
        "page_num": idx + 1,
        "sheet_name": sheet_name,
        "text": "\n".join(text_lines),
        "tables": [rows_data] if rows_data else [],
        "row_count": len(rows_data),
        "col_count": max((len(r) for r in rows_data), default=0),
    }


def _xlrd_cell_to_str(cell: Any, wb: Any) -> str:
    """Convert an xlrd Cell to a display string."""
    import xlrd

    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
            return dt.isoformat()
        except Exception:
            return str(cell.value)
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return "TRUE" if cell.value else "FALSE"
    if cell.ctype == xlrd.XL_CELL_ERROR:
        return "#ERR"
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        v = cell.value
        return str(int(v)) if v == int(v) else str(v)
    if cell.ctype == xlrd.XL_CELL_EMPTY:
        return ""
    return str(cell.value).strip()
