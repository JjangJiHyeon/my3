from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet


ROOT = Path(__file__).resolve().parent
DESKTOP = ROOT.parent


def _copy_cell_style(src, dst) -> None:
    if src.has_style:
        dst._style = copy(src._style)
    if src.font:
        dst.font = copy(src.font)
    if src.fill:
        dst.fill = copy(src.fill)
    if src.border:
        dst.border = copy(src.border)
    if src.alignment:
        dst.alignment = copy(src.alignment)
    if src.protection:
        dst.protection = copy(src.protection)
    if src.number_format:
        dst.number_format = src.number_format


def _clone_row_style(ws: Worksheet, source_row: int, target_row: int, col_start: int, col_end: int) -> None:
    for col in range(col_start, col_end + 1):
        _copy_cell_style(ws.cell(source_row, col), ws.cell(target_row, col))
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def _unmerge_if_exists(ws: Worksheet, cell_range: str) -> None:
    if cell_range in {str(rng) for rng in ws.merged_cells.ranges}:
        ws.unmerge_cells(cell_range)


def _merge_row_description(ws: Worksheet, row: int, start_col: str = "C", end_col: str = "F") -> None:
    ws.merge_cells(f"{start_col}{row}:{end_col}{row}")


def _set_wrap(ws: Worksheet, coord: str) -> None:
    cell = ws[coord]
    alignment = copy(cell.alignment) if cell.alignment else Alignment()
    alignment.wrap_text = True
    alignment.vertical = alignment.vertical or "center"
    cell.alignment = alignment


def _set_value(ws: Worksheet, coord: str, value: str) -> None:
    if isinstance(ws[coord], MergedCell):
        return
    ws[coord] = value
    _set_wrap(ws, coord)


def _update_design_workbook(path: Path) -> None:
    wb = load_workbook(path)

    ws = wb["기술 스택 및 미결사항"]
    for cell_range in [
        "E6:F11",
        "E12:F12",
        "E13:F13",
        "E14:F14",
        "E15:F15",
        "E16:F16",
        "B10:B11",
        "B15:B16",
    ]:
        _unmerge_if_exists(ws, cell_range)
    for row in range(6, 17):
        _merge_row_description(ws, row, "E", "F")

    tech_rows = {
        6: ("언어·환경", "Python 3.11.15, FastAPI, Uvicorn", "백엔드 API, 정적 파일 서빙, 업로드·삭제 job orchestration", "현재 로컬 실행 환경 기준 Python 3.11.15"),
        7: ("문서 파서", "PyMuPDF(fitz), pdfplumber, camelot-py, pymupdf4llm", "PDF 텍스트·레이아웃·표 파싱", "PDF 전용 파서"),
        8: ("", "python-docx, tika, olefile, openpyxl, xlrd", "DOC/DOCX/HWP/XLS/XLSX 파싱", "오피스 문서 파서"),
        9: ("OCR 보조", "easyocr, pytesseract, opencv-python, Pillow", "스캔형 PDF 및 이미지 텍스트 fallback OCR", "로컬 OCR fallback"),
        10: ("청킹", "chunking/strategies.py, chunking/builders.py", "text_first_with_visual_support, llm_ready_native 전략별 chunk 생성", "MAIN_STRATEGY / BASELINE_STRATEGY"),
        11: ("임베딩", "retrieval/vector_prep.py, retrieval/embedding_runner.py, numpy", "retrieval_text가 있는 chunk를 OpenAI 임베딩으로 변환", "OPENAI_API_KEY 필요"),
        12: ("임베딩 모델", "OpenAI text-embedding-3-large", "3072차원 문서 청크 임베딩", ".env.example 기준"),
        13: ("벡터 DB", "chromadb, langchain-chroma, document_chunks collection", "로컬 persistent collection 저장·검색", "chroma_indexes/latest.json 사용"),
        14: ("QA·요약 체인", "langchain, langchain-openai, ChatOpenAI(gpt-5.2)", "QA·문서 전체 요약 응답 생성", "scripts/run_rag_api.py에서 사용"),
        15: ("UI", "static/index.html, static/summary.html, FastAPI FileResponse", "문서 목록, 업로드, 삭제, QA, 요약, 품질 대시보드 UI", "app.py가 정적 화면 서빙"),
        16: ("관측/로그", "Python logging, latency_events.jsonl, LangSmith(optional)", "실행 로그 및 trace 기록", "LANGSMITH_TRACING=true 시만 trace"),
    }
    for row, (col_b, col_c, col_d, col_e) in tech_rows.items():
        _set_value(ws, f"B{row}", col_b)
        _set_value(ws, f"C{row}", col_c)
        _set_value(ws, f"D{row}", col_d)
        _set_value(ws, f"E{row}", col_e)
    ws.row_dimensions[6].height = 42
    for row in range(7, 17):
        ws.row_dimensions[row].height = 36

    ws = wb["프로그램 구성안"]
    _set_value(
        ws,
        "B5",
        "FastAPI 앱(app.py)이 문서 업로드/삭제, 문서 목록, 정적 UI, RAG proxy를 담당하고, "
        "업로드 job은 parse_document → write_root_llm_exports → scripts/build_chunks.py → "
        "scripts/run_retrieval_prep.py → scripts/run_chroma_ingest.py 순으로 최신 인덱스를 갱신한다.\n"
        "조회는 scripts/run_rag_api.py + rag_api/* 가 Chroma 검색과 QA/문서 전체 요약 생성을 담당한다.",
    )
    _set_value(
        ws,
        "C9",
        "실제 파일 기준으로 app.py, app_support, parsers, chunking, retrieval, chroma_store, scripts, rag_api, static으로 구성된다.\n"
        "아래 표는 업로드→파싱→청킹→임베딩→Chroma 적재→RAG QA/요약 워크플로우에 직접 관여하는 파일만 정리했다.",
    )

    for cell_range in [f"C{row}:F{row}" for row in range(13, 19)]:
        _unmerge_if_exists(ws, cell_range)

    file_rows = [
        ("app.py", "FastAPI 진입점. 문서 목록, 업로드/삭제 job, /api/rag/qa·/summary proxy, 정적 UI, 품질 리포트 엔드포인트를 제공한다."),
        (".env.example", "OpenAI chat/embedding 모델, Chroma collection/persist 경로, LangSmith 옵션 환경변수 예시를 정의한다."),
        ("requirements.txt", "FastAPI, 파서, OCR, OpenAI, Chroma, LangChain 등 실제 런타임 의존성을 정의한다."),
        ("app_support/artifact_cleanup.py", "생성 산출물 정리 유틸리티. 서버 시작 시 cleanup_generated_artifacts와 삭제 후속 정리에 사용된다."),
        ("app_support/export_to_gpt.py", "파싱 결과에서 *_llm_ready.json, *_llm_report.md용 payload를 만들고 write_root_llm_exports를 제공한다."),
        ("parsers/__init__.py", "PARSER_MAP, SUPPORTED_EXTENSIONS, parse_document 진입점을 제공해 확장자별 파서로 라우팅한다."),
        ("parsers/document_router.py", "PDF 초기 시그니처를 분석해 slide/report/dashboard 성격을 판단하고 page routing 근거를 계산한다."),
        ("parsers/pdf_parser.py", "PDF 페이지/블록/표/OCR/preview 이미지를 생성하고 라우팅 결과를 반영한 parse_pdf를 제공한다."),
        ("parsers/doc_parser.py", "DOC/DOCX 텍스트·테이블을 추출해 parse_doc 결과를 만든다. Tika, OLE 추출, win32 fallback 경로를 포함한다."),
        ("parsers/hwp_parser.py", "OLE 기반 HWP 본문/메타데이터를 해석해 parse_hwp 결과를 생성한다."),
        ("parsers/xlsx_parser.py", "openpyxl/xlrd로 XLSX/XLS 시트별 행 데이터를 읽어 page payload 형태로 변환한다."),
        ("parsers/ocr_utils.py", "EasyOCR/pytesseract 전처리, variant 생성, OCR 실행 및 결과 정리 유틸리티를 제공한다."),
        ("parsers/table_utils.py", "표 후보 감지, 병합, dashboard형 표 판별, 품질 점수 계산에 필요한 보조 함수를 제공한다."),
        ("parsers/quality_utils.py", "파싱 결과의 quality_score와 grade 계산 로직을 제공한다."),
        ("scripts/build_chunks.py", "업로드된 parsed_results를 읽어 문서별 chunk JSON과 chunk_summary.json을 생성하는 CLI 진입점이다."),
        ("chunking/builders.py", "build_all_for_document와 summarize_chunks로 전략별 chunk payload와 요약 통계를 조립한다."),
        ("chunking/io.py", "parsed_results/review/llm_ready 입력 로드와 chunk output 직렬화를 담당한다."),
        ("chunking/strategies.py", "text_first_with_visual_support, llm_ready_native 두 전략의 실제 chunk 생성 로직을 제공한다."),
        ("chunking/utils.py", "chunk id, block/page 메타, long text 분할, summary 보조 계산 등을 처리한다."),
        ("chunking/visual_structured.py", "표·차트 블록에서 retrieval 보조용 structured record를 만든다."),
        ("scripts/run_retrieval_prep.py", "chunks 디렉터리를 받아 retrieval.vector_prep.prepare_vectors를 실행하는 CLI 진입점이다."),
        ("retrieval/metadata_builder.py", "chunk JSON을 탐색해 임베딩 전 metadata record 목록을 만든다."),
        ("retrieval/metadata_schema.py", "metadata record 정규화, evidence preview, stable vector id 생성 규칙을 제공한다."),
        ("retrieval/embedding_runner.py", "OpenAI 임베딩 호출, API key 로드, 배치 임베딩 실행을 담당한다."),
        ("retrieval/vector_prep.py", "metadata JSONL, id_map, vectors.npz를 생성하고 vector_indexes/latest.json을 갱신한다."),
        ("scripts/run_chroma_ingest.py", "vector run 산출물을 받아 chroma_store.ingest.ingest_vector_run을 실행하는 CLI 진입점이다."),
        ("chroma_store/config.py", "Chroma persist 경로, collection 이름, metadata coercion과 project path 해석을 담당한다."),
        ("chroma_store/collection_builder.py", "persistent client 생성, collection 생성/조회, batch add_records 로직을 제공한다."),
        ("chroma_store/ingest.py", "vector run을 Chroma에 적재하고 latest.json, id_manifest, ingest summary를 갱신한다."),
        ("scripts/run_rag_api.py", "별도 FastAPI RAG 서버. /qa, /summary 엔드포인트와 run artifact 기록을 담당한다."),
        ("rag_api/config.py", "RAG API가 chroma_indexes/latest.json과 환경변수를 읽어 최신 Chroma 설정을 해석한다."),
        ("rag_api/retriever.py", "ChromaRetriever가 QA relevance 검색과 summary coverage 기반 representative 검색을 수행한다."),
        ("rag_api/summary_ranker.py", "summary 후보 chunk를 rerank_summary_documents로 재정렬한다."),
        ("rag_api/qa_chain.py", "QA 프롬프트와 ChatOpenAI 호출을 구성해 최종 답변을 생성한다."),
        ("rag_api/summary_chain.py", "문서 전체 요약 프롬프트와 ChatOpenAI 호출을 구성한다."),
        ("rag_api/source_formatter.py", "중복 제거된 document/page source 목록을 응답 포맷으로 정리한다."),
        ("rag_api/observability.py", "latency_events.jsonl 기록, stage_timer, optional LangSmith trace 유틸리티를 제공한다."),
        ("rag_api/schemas.py", "QARequest, SummaryRequest, RagResponse 등 FastAPI request/response 스키마를 정의한다."),
        ("static/index.html", "문서 목록, 업로드/삭제 polling, 문서 선택, QA, 요약 UI와 API 호출 로직을 포함한다."),
        ("static/summary.html", "품질 대시보드 화면과 prototype summary 조회 UI를 제공한다."),
    ]

    needed_rows = len(file_rows)
    existing_rows = 6  # rows 13..18
    extra_rows = max(0, needed_rows - existing_rows)
    if extra_rows:
        ws.insert_rows(19, amount=extra_rows)
    for row in range(13, 13 + needed_rows):
        _clone_row_style(ws, 13, row, 2, 6)
        _merge_row_description(ws, row)
    for row in range(13 + needed_rows, ws.max_row + 1):
        if f"C{row}:F{row}" in {str(rng) for rng in ws.merged_cells.ranges}:
            ws.unmerge_cells(f"C{row}:F{row}")

    for idx, (filename, description) in enumerate(file_rows, start=13):
        _set_value(ws, f"B{idx}", filename)
        _set_value(ws, f"C{idx}", description)
        ws.row_dimensions[idx].height = 34
    ws.row_dimensions[13].height = 44

    ws = wb["프로그램 워크플로우"]
    _set_value(
        ws,
        "C6",
        "1. 서버 시작 시 RESULT_DIR 캐시 warm-up과 선택적 generated artifact cleanup을 수행한다.\n"
        "2. 사용자가 static/index.html UI에서 문서를 업로드하면 app.py가 job_id를 만들고 백그라운드 thread에서 parse_document를 실행한다.\n"
        "3. 파싱 성공 문서는 parsed_results에 저장되고 app_support/export_to_gpt.py가 *_llm_ready.json, *_llm_report.md를 생성한다.\n"
        "4. 이어서 scripts/build_chunks.py → scripts/run_retrieval_prep.py → scripts/run_chroma_ingest.py 순으로 최신 chunk/vector/Chroma 인덱스를 갱신한다.\n"
        "5. 사용자는 /api/rag/documents 기반 문서 목록에서 대상을 선택하고 app.py의 /api/rag/qa 또는 /api/rag/summary를 호출한다.\n"
        "6. app.py는 요청을 scripts/run_rag_api.py의 /qa 또는 /summary로 프록시하고, RAG API는 rag_api/config.py에서 chroma_indexes/latest.json을 읽어 최신 컬렉션을 연다.\n"
        "7. rag_api/retriever.py가 filename_filter·strategy_name·top_k 기준으로 chunk를 조회하고, qa_chain.py 또는 summary_chain.py가 gpt-5.2로 최종 응답을 만든다.\n"
        "8. 응답은 sources(document,page)와 함께 UI에 표시되고 latency_events.jsonl 및 선택적 LangSmith trace에 기록된다.\n"
        "9. 삭제 요청 시 app.py가 관련 산출물을 stage 후 filtered vector run을 만든 뒤 scripts/run_chroma_ingest.py로 최신 인덱스를 재생성한다.",
    )
    _set_value(
        ws,
        "C10",
        "upload_rag_document → _process_upload_job → parse_document → write_root_llm_exports → "
        "scripts/build_chunks.py → scripts/run_retrieval_prep.py → scripts/run_chroma_ingest.py. "
        "진행률(progress/message)과 상태(queued/running/completed/failed)를 polling으로 조회한다.",
    )
    _set_value(
        ws,
        "C11",
        "app.py /api/rag/qa, /api/rag/summary → _rag_api_post → scripts/run_rag_api.py FastAPI. "
        "QA는 query+filename_filter+top_k, Summary는 filename+strategy_name+top_k 기반이며 응답은 sources(document,page)를 포함한다.",
    )
    _set_value(
        ws,
        "C12",
        "delete_rag_document → _process_delete_job으로 문서와 managed artifact를 안전 staging한 뒤 "
        "filtered vector run 기준으로 Chroma 최신 인덱스를 다시 만든다. "
        "/prototype-summary 화면에서는 quality score, failure source, report markdown을 조회한다.",
    )
    for row in (6, 10, 11, 12):
        ws.row_dimensions[row].height = 70 if row == 6 else 44

    wb.save(path)


def _update_feature_workbook(path: Path) -> None:
    wb = load_workbook(path)

    ws = wb["요건 정의 상세"]
    _set_value(
        ws,
        "F5",
        "static/index.html upload accept, app.py _safe_upload_name(), parsers/__init__.py SUPPORTED_EXTENSIONS",
    )
    _set_value(
        ws,
        "F6",
        "parsers/__init__.py parse_document, parsers/pdf_parser.py parse_pdf, "
        "parsers/doc_parser.py parse_doc, parsers/hwp_parser.py parse_hwp, parsers/xlsx_parser.py parse_excel",
    )
    _set_value(
        ws,
        "F7",
        "parsers/pdf_parser.py, parsers/doc_parser.py, parsers/hwp_parser.py, parsers/xlsx_parser.py",
    )
    _set_value(
        ws,
        "F8",
        "parsers/quality_utils.py calculate_quality_score, parsed_results/*.json",
    )
    _set_value(
        ws,
        "F9",
        "chunking/strategies.py, chunking/builders.py, scripts/build_chunks.py",
    )
    _set_value(
        ws,
        "F10",
        "chunks/{doc_id}.chunk_summary.json",
    )
    _set_value(
        ws,
        "F11",
        "scripts/run_retrieval_prep.py, retrieval/vector_prep.py prepare_vectors, "
        "retrieval/embedding_runner.py embed_texts, text-embedding-3-large(3072)",
    )
    _set_value(
        ws,
        "F12",
        "scripts/run_chroma_ingest.py, chroma_store/ingest.py ingest_vector_run, "
        "chroma_indexes/latest.json",
    )
    _set_value(
        ws,
        "F13",
        "app.py _process_upload_job/_process_delete_job, scripts/build_chunks.py, "
        "scripts/run_retrieval_prep.py, scripts/run_chroma_ingest.py",
    )
    _set_value(
        ws,
        "F14",
        "POST /api/rag/qa → app.py proxy → scripts/run_rag_api.py /qa → "
        "rag_api/retriever.py retrieve → rag_api/qa_chain.py run_qa_chain",
    )
    _set_value(
        ws,
        "F15",
        "POST /api/rag/summary → app.py proxy → scripts/run_rag_api.py /summary → "
        "rag_api/retriever.py retrieve_representative → rag_api/summary_chain.py run_summary_chain",
    )
    _set_value(
        ws,
        "F16",
        "rag_api/retriever.py summary_coverage_cache / summary_selection_cache",
    )
    _set_value(
        ws,
        "F17",
        "rag_api/source_formatter.py, app.py _is_document_title_query(), _document_title_response()",
    )
    _set_value(
        ws,
        "F18",
        "rag_api/config.py, rag_api/observability.py stage_timer/optional_traceable",
    )

    ws = wb["Input 및 Output 정의"]
    _set_value(
        ws,
        "C5",
        "• 입력 문서\n"
        " - 지원 형식: PDF, DOC, DOCX, HWP, XLS, XLSX\n"
        " - 현재 documents/ 코퍼스: PDF 7건, DOC 1건, HWP 1건, XLSX 1건\n"
        " - 현재 XLSX 문서: 2026년도 IT재고실사_유지현.xlsx\n"
        "• 입력 파라미터\n"
        " - QA: query, strategy_name, top_k, filename_filter\n"
        " - Summary: filename, strategy_name, top_k\n"
        "• 운영 입력\n"
        " - .env 기반 OPENAI_API_KEY, 선택적 LANGSMITH_* 설정, chroma_indexes/latest.json",
    )
    ws.row_dimensions[5].height = 120

    wb.save(path)


def update_workbooks(design_path: Path, feature_path: Path) -> None:
    _update_design_workbook(design_path)
    _update_feature_workbook(feature_path)


def main() -> None:
    design_path = DESKTOP / "개발설계서_my3_v0.1.xlsx"
    feature_path = DESKTOP / "기능정의서_my3_v0.1.xlsx"
    update_workbooks(design_path, feature_path)
    print(f"Updated: {design_path}")
    print(f"Updated: {feature_path}")


if __name__ == "__main__":
    main()
